"""File I/O utilities: SHA-256, stat, metadata extraction."""

from __future__ import annotations

import asyncio
import hashlib
import os
from dataclasses import dataclass
from pathlib import Path

from ...filetypes import OFFICE_EXTENSIONS
from ...utils.pdf_document import PDFDocument

# Optional office doc support
try:
    from docx import Document
except ImportError:
    Document = None  # type: ignore[assignment]
try:
    from pptx import Presentation  # type: ignore[reportMissingImports]
except ImportError:
    Presentation = None  # type: ignore[assignment]
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[assignment]

# Metadata truncation limits
TRUNC_TITLE = 500
TRUNC_AUTHOR = 200
TRUNC_SUBJECT = 1000
TRUNC_KEYWORDS = 1000


class MetadataExtractionError(Exception):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


# ── SHA-256 ────────────────────────────────────────────────────────


def _sha256_sync(filepath: str) -> str:
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        while chunk := f.read(65536):
            h.update(chunk)
    return h.hexdigest()


async def compute_sha256(filepath: str) -> str:
    return await asyncio.to_thread(_sha256_sync, filepath)


# ── file stat ──────────────────────────────────────────────────────


@dataclass(frozen=True)
class FileStat:
    size_bytes: int
    mtime_ms: int


async def get_file_stat(filepath: str) -> FileStat:
    def _stat() -> FileStat:
        st = os.stat(filepath)
        return FileStat(size_bytes=st.st_size, mtime_ms=int(st.st_mtime * 1000))

    return await asyncio.to_thread(_stat)


# ── metadata extraction ────────────────────────────────────────────


async def extract_metadata(filepath: str) -> dict:
    """Extract metadata from file.  Returns dict with keys matching docs table columns.
    All string fields are protectively truncated."""
    ext = Path(filepath).suffix.lower().lstrip(".")

    result = {
        "page_count": None,
        "title": None,
        "author": None,
        "subject": None,
        "keywords": None,
        "is_image_based": 0,
    }

    if ext == "pdf":
        await _extract_pdf_meta(filepath, result)

    elif ext in OFFICE_EXTENSIONS:
        await _extract_office_meta(filepath, ext, result)

    # truncate all string fields
    for field, limit in [
        ("title", TRUNC_TITLE),
        ("author", TRUNC_AUTHOR),
        ("subject", TRUNC_SUBJECT),
        ("keywords", TRUNC_KEYWORDS),
    ]:
        val = result.get(field)
        if val and len(val) > limit:
            result[field] = val[:limit]

    return result


# ── PDF metadata ───────────────────────────────────────────────────


async def _extract_pdf_meta(filepath: str, result: dict) -> None:
    def _extract() -> None:
        pdf_doc = None
        try:
            try:
                pdf_doc = PDFDocument(filepath)
                result["page_count"] = pdf_doc.page_count
            except Exception as exc:
                raise MetadataExtractionError("open_failed", str(exc) or "Failed to open document") from exc

            try:
                meta = pdf_doc.metadata
            except Exception as exc:
                raise MetadataExtractionError("read_metadata_failed", str(exc) or "Failed to read document metadata") from exc

            try:
                result["title"] = meta.get("Title") or None
                result["author"] = meta.get("Author") or None
                result["subject"] = meta.get("Subject") or None
                result["keywords"] = meta.get("Keywords") or None
            except Exception as exc:
                raise MetadataExtractionError("read_metadata_failed", str(exc) or "Failed to read document metadata") from exc
        finally:
            if pdf_doc is not None:
                pdf_doc.close()

    await asyncio.to_thread(_extract)


# ── Office metadata ────────────────────────────────────────────────


async def _extract_office_meta(filepath: str, ext: str, result: dict) -> None:
    def _extract() -> None:
        if ext == "docx":
            if Document is None:
                return
            try:
                doc = Document(filepath)
            except Exception as exc:
                raise MetadataExtractionError("open_failed", str(exc) or "Failed to open document") from exc

            try:
                cp = doc.core_properties
                result["title"] = cp.title or None
                result["author"] = cp.author or None
                result["subject"] = cp.subject or None
                result["keywords"] = cp.keywords or None
            except Exception as exc:
                raise MetadataExtractionError("read_metadata_failed", str(exc) or "Failed to read document metadata") from exc

        elif ext == "pptx":
            if Presentation is None:
                return
            try:
                prs = Presentation(filepath)
                result["page_count"] = len(prs.slides)
            except Exception as exc:
                raise MetadataExtractionError("open_failed", str(exc) or "Failed to open document") from exc

            try:
                cp = prs.core_properties
                result["title"] = cp.title or None
                result["author"] = cp.author or None
                result["subject"] = cp.subject or None
                result["keywords"] = cp.keywords or None
            except Exception as exc:
                raise MetadataExtractionError("read_metadata_failed", str(exc) or "Failed to read document metadata") from exc

        elif ext == "xlsx":
            if load_workbook is None:
                return
            wb = None
            try:
                wb = load_workbook(filepath, read_only=True)
                result["page_count"] = len(wb.sheetnames)
            except Exception as exc:
                raise MetadataExtractionError("open_failed", str(exc) or "Failed to open document") from exc
            try:
                cp = wb.properties
                result["title"] = cp.title or None
                result["author"] = cp.creator or None
                result["subject"] = cp.subject or None
                result["keywords"] = cp.keywords or None
            except Exception as exc:
                raise MetadataExtractionError("read_metadata_failed", str(exc) or "Failed to read document metadata") from exc
            finally:
                if wb is not None:
                    wb.close()

    await asyncio.to_thread(_extract)
