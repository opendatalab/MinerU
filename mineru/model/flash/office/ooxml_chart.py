# Copyright (c) Opendatalab. All rights reserved.
"""Flash Office 文档中的 OOXML 图表解析与表格化渲染。"""

from __future__ import annotations

import math
from dataclasses import dataclass, field
from datetime import date, datetime, time
from html import escape
from io import BytesIO
from typing import Any, Final

from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils.cell import range_to_tuple
from openpyxl.utils.datetime import MAC_EPOCH, WINDOWS_EPOCH, from_excel
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

_CHART_NS: Final = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_DRAWING_NS: Final = "http://schemas.openxmlformats.org/drawingml/2006/main"
_NS: Final = {"c": _CHART_NS, "a": _DRAWING_NS}
_MAX_CACHE_INDEX_SPAN: Final = 100_000
_PLOT_TAGS: Final = (
    "areaChart",
    "area3DChart",
    "barChart",
    "bar3DChart",
    "bubbleChart",
    "doughnutChart",
    "lineChart",
    "line3DChart",
    "ofPieChart",
    "pieChart",
    "pie3DChart",
    "radarChart",
    "scatterChart",
    "stockChart",
    "surfaceChart",
    "surface3DChart",
)


@dataclass
class SeriesSpec:
    """保存单个 OOXML 图表序列的公式引用、名称和缓存数据。"""

    name_formula: str | None = None
    literal_name: str | None = None
    cat_formula: str | None = None
    x_formula: str | None = None
    val_formula: str | None = None
    y_formula: str | None = None
    bubble_size_formula: str | None = None
    cached_categories: list[str] = field(default_factory=list)
    cached_x_values: list[str] = field(default_factory=list)
    cached_values: list[str] = field(default_factory=list)
    cached_bubble_sizes: list[str] = field(default_factory=list)


@dataclass
class ChartSpec:
    """保存图表类型、坐标轴信息和全部序列的规范化描述。"""

    chart_type: str
    plot_kind: str
    title: str = ""
    category_axis_title: str = ""
    value_axis_title: str = ""
    x_axis_title: str = ""
    has_date_axis: bool = False
    date_1904: bool = False
    series: list[SeriesSpec] = field(default_factory=list)


def html_table_from_excel_bytes(excel_bytes: bytes) -> str:
    """将嵌入工作簿中第一个非空工作表转换为 HTML 表格。"""
    if not excel_bytes:
        return ""

    try:
        workbook = load_workbook(
            filename=BytesIO(excel_bytes),
            data_only=True,
            read_only=True,
        )
    except Exception:
        return ""

    try:
        for worksheet in workbook.worksheets:
            rows = _extract_non_empty_worksheet_rows(worksheet)
            if rows:
                return _render_embedded_workbook_table(rows)
    finally:
        workbook.close()

    return ""


def _extract_non_empty_worksheet_rows(worksheet: Worksheet) -> list[list[str]]:
    """提取工作表中首尾有内容的行，避免空 sheet 或尾部空列撑大兜底表格。"""
    raw_rows: list[list[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        stringified_row = [_stringify_cell_value(value) for value in row]
        raw_rows.append(_trim_trailing_empty_values(stringified_row))

    first_row_idx = _find_first_non_empty_row(raw_rows)
    if first_row_idx is None:
        return []

    last_row_idx = _find_last_non_empty_row(raw_rows)
    if last_row_idx is None:
        return []

    last_row_exclusive = last_row_idx + 1
    rows = raw_rows[first_row_idx:last_row_exclusive]
    width = max(len(row) for row in rows)
    return [row + [""] * (width - len(row)) for row in rows]


def _trim_trailing_empty_values(values: list[str]) -> list[str]:
    """移除行尾空值，保留中间空单元格的位置。"""
    end = len(values)
    while end > 0 and values[end - 1] == "":
        end -= 1
    return values[:end]


def _find_first_non_empty_row(rows: list[list[str]]) -> int | None:
    """返回第一行非空行的索引，未找到时返回 None。"""
    for idx, row in enumerate(rows):
        if any(value != "" for value in row):
            return idx
    return None


def _find_last_non_empty_row(rows: list[list[str]]) -> int | None:
    """返回最后一行非空行的索引。"""
    for idx in range(len(rows) - 1, -1, -1):
        if any(value != "" for value in rows[idx]):
            return idx
    return None


def _render_embedded_workbook_table(rows: list[list[str]]) -> str:
    """将嵌入 workbook 的二维数据渲染为紧凑 HTML 表格，首行作为表头。"""
    if not rows:
        return ""

    headers = rows[0]
    data_rows = rows[1:]
    html_parts = ["<table><thead><tr>"]
    for header in headers:
        html_parts.append(f"<th>{escape(header)}</th>")
    html_parts.append("</tr></thead><tbody>")

    for row in data_rows:
        html_parts.append("<tr>")
        for value in row:
            html_parts.append(f"<td>{escape(value)}</td>")
        html_parts.append("</tr>")

    html_parts.append("</tbody></table>")
    return "".join(html_parts)


def extract_chart_html_from_ooxml(chart_xml: bytes, workbook_bytes: bytes | None) -> str:
    """解析 OOXML 图表并按工作簿、缓存和普通表格的优先级生成 HTML。"""
    spec = parse_chart_spec_from_ooxml(chart_xml)
    if spec is None or not spec.series:
        if workbook_bytes:
            return html_table_from_excel_bytes(workbook_bytes)
        return ""

    chart_cache_html = render_chart_html_from_cache(spec)

    if workbook_bytes:
        chart_html = render_chart_html_from_workbook(spec, workbook_bytes)
        if chart_html:
            return chart_html
        if chart_cache_html:
            return chart_cache_html
        workbook_table_html = html_table_from_excel_bytes(workbook_bytes)
        if workbook_table_html:
            return workbook_table_html

    return chart_cache_html


def parse_chart_spec_from_ooxml(chart_xml: bytes) -> ChartSpec | None:
    """解析图表 XML，生成统一的图表类型、坐标轴和序列描述。"""
    try:
        root = etree.fromstring(
            chart_xml,
            parser=etree.XMLParser(
                load_dtd=False,
                resolve_entities=False,
                no_network=True,
            ),
        )
    except (etree.XMLSyntaxError, TypeError, ValueError):
        return None

    plot_area = root.find(".//c:plotArea", namespaces=_NS)
    if plot_area is None:
        return None

    plot_elements = _collect_plot_elements(plot_area)
    if not plot_elements:
        return None

    has_date_axis = plot_area.find("c:dateAx", namespaces=_NS) is not None
    plot_kinds = {_plot_kind_from_tag_name(tag_name, has_date_axis) for tag_name, _ in plot_elements}
    if plot_kinds == {"scatter"}:
        plot_kind = "scatter"
    elif plot_kinds == {"bubble"}:
        plot_kind = "bubble"
    elif plot_kinds <= {"category", "date"}:
        plot_kind = "date" if "date" in plot_kinds else "category"
    else:
        return None

    category_axis_title = ""
    axis = plot_area.find("c:dateAx", namespaces=_NS)
    if axis is None:
        axis = plot_area.find("c:catAx", namespaces=_NS)
    if axis is not None:
        category_axis_title = _extract_title_text(axis.find("c:title", namespaces=_NS))  # type: ignore

    x_axis_title = ""
    value_axis_title = ""
    if plot_kind in {"scatter", "bubble"}:
        for axis in plot_area.findall("c:valAx", namespaces=_NS):
            axis_pos = axis.find("c:axPos", namespaces=_NS)
            axis_position = axis_pos.get("val") if axis_pos is not None else ""
            title = _extract_title_text(axis.find("c:title", namespaces=_NS))  # type: ignore
            if axis_position == "b" and not x_axis_title:
                x_axis_title = title
            elif axis_position == "l" and not value_axis_title:
                value_axis_title = title
        if not x_axis_title:
            x_axis_title = category_axis_title
    else:
        axis = plot_area.find("c:valAx", namespaces=_NS)
        if axis is not None:
            value_axis_title = _extract_title_text(axis.find("c:title", namespaces=_NS))  # type: ignore

    series_specs = []
    for _, plot_element in plot_elements:
        for series_element in plot_element.findall("c:ser", namespaces=_NS):
            series_specs.append(
                SeriesSpec(
                    name_formula=_extract_tx_formula(series_element.find("c:tx", namespaces=_NS)),  # type: ignore
                    literal_name=_extract_tx_text(series_element.find("c:tx", namespaces=_NS)),  # type: ignore
                    cat_formula=_extract_reference_formula(series_element.find("c:cat", namespaces=_NS)),  # type: ignore
                    x_formula=_extract_reference_formula(series_element.find("c:xVal", namespaces=_NS)),  # type: ignore
                    val_formula=_extract_reference_formula(series_element.find("c:val", namespaces=_NS)),  # type: ignore
                    y_formula=_extract_reference_formula(series_element.find("c:yVal", namespaces=_NS)),  # type: ignore
                    bubble_size_formula=_extract_reference_formula(series_element.find("c:bubbleSize", namespaces=_NS)),  # type: ignore
                    cached_categories=_extract_reference_cache(
                        series_element.find("c:cat", namespaces=_NS),  # type: ignore
                        date_hint=has_date_axis,
                        date_1904=_chart_uses_date_1904(root),
                    ),
                    cached_x_values=_extract_reference_cache(series_element.find("c:xVal", namespaces=_NS)),  # type: ignore
                    cached_values=_extract_reference_cache(
                        _first_non_none(
                            series_element.find("c:val", namespaces=_NS),
                            series_element.find("c:yVal", namespaces=_NS),
                        )  # type: ignore
                    ),
                    cached_bubble_sizes=_extract_reference_cache(series_element.find("c:bubbleSize", namespaces=_NS)),  # type: ignore
                )
            )

    return ChartSpec(
        chart_type=(plot_elements[0][0] if len(plot_elements) == 1 else "comboChart"),
        plot_kind=plot_kind,
        title=_extract_title_text(root.find(".//c:chart/c:title", namespaces=_NS)),  # type: ignore
        category_axis_title=category_axis_title,
        value_axis_title=value_axis_title,
        x_axis_title=x_axis_title,
        has_date_axis=has_date_axis,
        date_1904=_chart_uses_date_1904(root),
        series=series_specs,
    )


def render_chart_html_from_workbook(spec: ChartSpec, workbook_bytes: bytes) -> str:
    """根据图表公式从嵌入工作簿读取数据并渲染 HTML 表格。"""
    try:
        workbook = load_workbook(
            filename=BytesIO(workbook_bytes),
            data_only=True,
            read_only=True,
        )
    except Exception:
        return ""

    try:
        if spec.plot_kind in {"category", "date"}:
            return _render_category_like_chart_from_workbook(spec, workbook)
        if spec.plot_kind == "scatter":
            return _render_scatter_like_chart_from_workbook(spec, workbook)
        if spec.plot_kind == "bubble":
            return _render_bubble_chart_from_workbook(spec, workbook)
        return ""
    finally:
        workbook.close()


def render_chart_html_from_cache(spec: ChartSpec) -> str:
    """在工作簿不可用时使用 OOXML 图表缓存数据渲染 HTML 表格。"""
    if spec.plot_kind in {"category", "date"}:
        categories = []
        for series in spec.series:
            if series.cached_categories:
                categories = series.cached_categories
                break

        series_names = []
        series_values = []
        for idx, series in enumerate(spec.series, start=1):
            series_names.append(_resolve_series_name(series, idx))
            series_values.append(series.cached_values)

        row_count = max(
            len(categories),
            max((len(values) for values in series_values), default=0),
        )
        if not series_names or row_count == 0:
            return ""

        headers = [spec.category_axis_title or ""] + series_names
        columns = [categories] + series_values
        return _render_html_table(headers, columns, row_count)

    if spec.plot_kind == "scatter":
        return _render_scatter_like_chart_from_cache(spec)

    if spec.plot_kind == "bubble":
        return _render_bubble_chart_from_cache(spec)

    return ""


def _render_category_like_chart_from_workbook(spec: ChartSpec, workbook: Workbook) -> str:
    """从工作簿渲染分类轴或日期轴图表的二维 HTML 表格。"""
    categories = []

    for series in spec.series:
        if not series.cat_formula:
            continue
        read_result = _read_formula_vector(workbook, series.cat_formula)
        if read_result is None:
            return ""
        _, values = read_result
        categories = values
        break

    series_names = []
    series_values = []
    for idx, series in enumerate(spec.series, start=1):
        if not series.val_formula:
            return ""
        read_result = _read_formula_vector(workbook, series.val_formula)
        if read_result is None:
            return ""
        _, values = read_result
        series_names.append(_resolve_series_name(series, idx, workbook))
        series_values.append(values)

    row_count = max(
        len(categories),
        max((len(values) for values in series_values), default=0),
    )
    if not series_names or row_count == 0:
        return ""

    headers = [spec.category_axis_title or ""] + series_names
    columns = [
        _stringify_series_values(
            categories,
            date_hint=spec.has_date_axis,
            date_1904=spec.date_1904,
        )
    ]
    columns.extend(_stringify_series_values(values) for values in series_values)
    return _render_html_table(headers, columns, row_count)


def _render_scatter_like_chart_from_workbook(spec: ChartSpec, workbook: Workbook) -> str:
    """读取工作簿中的散点图横纵轴数据并渲染 HTML 表格。"""
    x_sequences, series_names, series_y_values = _read_scatter_axes_from_workbook(
        spec,
        workbook,
    )
    return _render_scatter_like_chart_table(
        x_sequences,
        series_names,
        series_y_values,
        x_axis_title=spec.x_axis_title,
    )


def _render_bubble_chart_from_workbook(spec: ChartSpec, workbook: Workbook) -> str:
    """读取工作簿中的气泡图横纵轴和尺寸数据并渲染 HTML 表格。"""
    x_sequences, series_names, series_y_values, series_sizes = _read_bubble_axes_from_workbook(
        spec,
        workbook,
    )
    return _render_bubble_chart_table(
        x_sequences,
        series_names,
        series_y_values,
        series_sizes,
        x_axis_title=spec.x_axis_title,
    )


def _render_scatter_like_chart_from_cache(spec: ChartSpec) -> str:
    """使用 OOXML 缓存的横纵轴数据渲染散点图 HTML 表格。"""
    x_sequences = []
    series_names = []
    series_y_values = []
    for idx, series in enumerate(spec.series, start=1):
        if not series.cached_x_values or not series.cached_values:
            return ""
        x_sequences.append(series.cached_x_values)
        series_names.append(_resolve_series_name(series, idx))
        series_y_values.append(series.cached_values)

    return _render_scatter_like_chart_table(
        x_sequences,
        series_names,
        series_y_values,
        x_axis_title=spec.x_axis_title,
    )


def _render_bubble_chart_from_cache(spec: ChartSpec) -> str:
    """使用 OOXML 缓存的横纵轴和尺寸数据渲染气泡图 HTML 表格。"""
    x_sequences = []
    series_names = []
    series_y_values = []
    series_sizes = []
    for idx, series in enumerate(spec.series, start=1):
        if not series.cached_x_values or not series.cached_values or not series.cached_bubble_sizes:
            return ""
        x_sequences.append(series.cached_x_values)
        series_names.append(_resolve_series_name(series, idx))
        series_y_values.append(series.cached_values)
        series_sizes.append(series.cached_bubble_sizes)

    return _render_bubble_chart_table(
        x_sequences,
        series_names,
        series_y_values,
        series_sizes,
        x_axis_title=spec.x_axis_title,
    )


def _read_scatter_axes_from_workbook(
    spec: ChartSpec, workbook: Workbook
) -> tuple[list[list[float]] | None, list[str], list[list[float]]]:
    """按序列公式读取散点图的 X 轴、序列名称和 Y 轴数据。"""
    x_sequences = []
    series_names = []
    series_y_values = []

    for idx, series in enumerate(spec.series, start=1):
        if not series.x_formula or not series.y_formula:
            return None, [], []

        x_read = _read_formula_vector(workbook, series.x_formula)
        if x_read is None:
            return None, [], []
        _, x_values = x_read
        x_sequences.append(x_values)

        y_read = _read_formula_vector(workbook, series.y_formula)
        if y_read is None:
            return None, [], []
        _, y_values = y_read
        series_names.append(_resolve_series_name(series, idx, workbook))
        series_y_values.append(y_values)

    return x_sequences, series_names, series_y_values


def _read_bubble_axes_from_workbook(
    spec: ChartSpec, workbook: Workbook
) -> tuple[list[list[float]] | None, list[str], list[list[float]], list[list[float]]]:
    """按序列公式读取气泡图的 X/Y 轴、名称和气泡尺寸数据。"""
    x_sequences = []
    series_names = []
    series_y_values = []
    series_sizes = []

    for idx, series in enumerate(spec.series, start=1):
        if not series.x_formula or not series.y_formula or not series.bubble_size_formula:
            return None, [], [], []

        x_read = _read_formula_vector(workbook, series.x_formula)
        if x_read is None:
            return None, [], [], []
        _, x_values = x_read
        x_sequences.append(x_values)

        y_read = _read_formula_vector(workbook, series.y_formula)
        bubble_size_read = _read_formula_vector(workbook, series.bubble_size_formula)
        if y_read is None or bubble_size_read is None:
            return None, [], [], []

        series_names.append(_resolve_series_name(series, idx, workbook))
        series_y_values.append(y_read[1])
        series_sizes.append(bubble_size_read[1])

    return x_sequences, series_names, series_y_values, series_sizes


def _read_formula_vector(workbook: Workbook, formula: str) -> tuple[str, list[Any]] | None:
    """解析单行或单列单元格公式，并从工作簿读取对应的一维数据。"""
    parsed = _parse_formula(formula)
    if parsed is None:
        return None

    sheet_name, min_col, min_row, max_col, max_row = parsed

    try:
        worksheet = workbook[sheet_name]
    except KeyError:
        return None

    if min_col != max_col and min_row != max_row:
        return None

    values = []
    if min_col == max_col:
        for row_idx in range(min_row, max_row + 1):
            values.append(worksheet.cell(row=row_idx, column=min_col).value)
    else:
        for col_idx in range(min_col, max_col + 1):
            values.append(worksheet.cell(row=min_row, column=col_idx).value)

    return sheet_name, values


def _read_formula_scalar(workbook: Workbook, formula: str) -> str | None:
    """读取公式引用区域的首个有效值并转换为字符串。"""
    read_result = _read_formula_vector(workbook, formula)
    if read_result is None:
        return None

    _, values = read_result
    if not values:
        return None

    value = values[0]
    if value in (None, ""):
        return None
    return _stringify_cell_value(value)


def _parse_formula(formula: str) -> tuple[str, int, int, int, int] | None:
    """将工作表区域公式解析为工作表名称和规范化单元格边界。"""
    formula = formula.strip()
    if not formula:
        return None
    if formula.startswith("="):
        formula = formula[1:]

    try:
        sheet_name, bounds = range_to_tuple(formula)
    except ValueError:
        return None

    if not all(isinstance(bound, int) for bound in bounds):
        return None

    min_col, min_row, max_col, max_row = bounds
    return _unescape_formula_sheet_name(sheet_name), min_col, min_row, max_col, max_row


def _unescape_formula_sheet_name(sheet_name: str) -> str:
    """还原 OOXML 公式中以双单引号转义的工作表名称。"""
    return sheet_name.replace("''", "'")


def _extract_reference_formula(container: etree._Element) -> str | None:
    """从字符串、数字或多级字符串引用容器中提取公式文本。"""
    ref_element = _find_reference_element(container)
    if ref_element is None:
        return None
    formula_element = ref_element.find("c:f", namespaces=_NS)
    if formula_element is None or formula_element.text is None:
        return None
    return formula_element.text.strip()


def _extract_reference_cache(
    container: etree._Element,
    *,
    date_hint: bool = False,
    date_1904: bool = False,
) -> list[str]:
    """从图表引用容器提取缓存数据，并按日期提示规范化值。"""
    ref_element = _find_reference_element(container)
    if ref_element is None:
        return []

    tag_name = etree.QName(ref_element).localname
    if tag_name == "multiLvlStrRef":
        return _extract_multilevel_string_cache(ref_element)

    cache_element = ref_element.find("c:strCache", namespaces=_NS)
    if cache_element is None:
        cache_element = ref_element.find("c:numCache", namespaces=_NS)
    if cache_element is None:
        return []

    return _extract_cache_points(
        cache_element,
        date_hint=date_hint,
        date_1904=date_1904,
    )


def _extract_cache_points(
    cache_element: etree._Element,
    *,
    date_hint: bool = False,
    date_1904: bool = False,
) -> list[str]:
    """按缓存点索引还原连续序列，并拒绝异常大的稀疏索引范围。"""
    points = {}
    for point in cache_element.findall("c:pt", namespaces=_NS):
        raw_index = point.get("idx")
        if raw_index is None:
            continue
        try:
            point_index = int(raw_index)
        except ValueError:
            continue

        value_element = point.find("c:v", namespaces=_NS)
        raw_value = value_element.text if value_element is not None else ""
        points[point_index] = _stringify_cache_value(
            raw_value,
            date_hint=date_hint,
            date_1904=date_1904,
        )

    if not points:
        return []

    max_index = max(points.keys())
    if max_index + 1 > _MAX_CACHE_INDEX_SPAN:
        return []

    return [points.get(index, "") for index in range(max_index + 1)]


def _extract_multilevel_string_cache(ref_element: etree._Element) -> list[str]:
    """合并多级分类缓存的同索引文本，生成扁平分类标签序列。"""
    level_maps = []
    max_index = -1
    for level in ref_element.findall("c:multiLvlStrCache/c:lvl", namespaces=_NS):
        values = {}
        for point in level.findall("c:pt", namespaces=_NS):
            raw_index = point.get("idx")
            if raw_index is None:
                continue
            try:
                point_index = int(raw_index)
            except ValueError:
                continue
            value_element = point.find("c:v", namespaces=_NS)
            values[point_index] = value_element.text if value_element is not None else ""
            max_index = max(max_index, point_index)
        level_maps.append(values)

    if max_index < 0:
        return []

    if max_index + 1 > _MAX_CACHE_INDEX_SPAN:
        return []

    rows = []
    for point_index in range(max_index + 1):
        parts = [value_map[point_index] for value_map in level_maps if value_map.get(point_index)]
        rows.append(" / ".join(parts))
    return rows


def _extract_tx_formula(tx_element: etree._Element) -> str | None:
    """从图表序列名称节点提取字符串引用公式。"""
    if tx_element is None:
        return None
    str_ref = tx_element.find("c:strRef", namespaces=_NS)
    if str_ref is None:
        return None
    formula_element = str_ref.find("c:f", namespaces=_NS)
    if formula_element is None or formula_element.text is None:
        return None
    return formula_element.text.strip()


def _extract_tx_text(tx_element: etree._Element) -> str | None:
    """从图表序列名称节点提取缓存文本或直接文本。"""
    if tx_element is None:
        return None

    str_cache = tx_element.find("c:strRef/c:strCache", namespaces=_NS)
    if str_cache is not None:
        values = _extract_cache_points(str_cache)
        return values[0] if values else None

    value_element = tx_element.find("c:v", namespaces=_NS)
    if value_element is not None and value_element.text:
        return value_element.text.strip()

    return None


def _extract_title_text(title_element: etree._Element) -> str:
    """拼接图表标题节点中的全部富文本片段。"""
    if title_element is None:
        return ""
    texts = title_element.findall(".//a:t", namespaces=_NS)
    return "".join(text.text or "" for text in texts).strip()


def _find_reference_element(container: etree._Element) -> object | None:
    """在容器中查找受支持的字符串、数字或多级字符串引用节点。"""
    if container is None:
        return None
    for tag_name in ("strRef", "numRef", "multiLvlStrRef"):
        ref_element = container.find(f"c:{tag_name}", namespaces=_NS)
        if ref_element is not None:
            return ref_element
    return None


def _first_non_none(*values: Any) -> object | None:
    """返回参数序列中的第一个非空对象。"""
    for value in values:
        if value is not None:
            return value
    return None


def _collect_plot_elements(plot_area: etree._Element) -> list[tuple[str, etree._Element]]:
    """收集绘图区中受支持的图表节点及其局部标签名。"""
    plot_elements = []
    for child in plot_area:
        if not isinstance(child.tag, str):
            continue
        tag_name = etree.QName(child).localname
        if tag_name in _PLOT_TAGS:
            plot_elements.append((tag_name, child))
    return plot_elements


def _plot_kind_from_tag_name(tag_name: str, has_date_axis: bool) -> str:
    """根据 OOXML 图表标签和日期轴信息归一化绘图类别。"""
    if tag_name == "scatterChart":
        return "scatter"
    if tag_name == "bubbleChart":
        return "bubble"
    if has_date_axis:
        return "date"
    return "category"


def _chart_uses_date_1904(root: etree._Element) -> bool:
    """判断图表是否使用以 1904 年为起点的 Excel 日期系统。"""
    date_1904 = root.find("c:date1904", namespaces=_NS)
    if date_1904 is None:
        return False
    return date_1904.get("val") == "1"


def _resolve_series_name(series: SeriesSpec, index: int, workbook: Workbook | None = None) -> str:
    """按工作簿引用、缓存名称和默认序号依次解析序列名称。"""
    if workbook is not None and series.name_formula:
        workbook_name = _read_formula_scalar(workbook, series.name_formula)
        if workbook_name:
            return workbook_name
    if series.literal_name:
        return series.literal_name
    return f"Series{index}"


def _get_shared_axis_values(sequences: list[list[Any]]) -> list[Any] | None:
    """判断多个序列是否共享相同轴值，并在一致时返回首个序列。"""
    if not sequences:
        return None

    normalized = [_normalize_sequence(sequence) for sequence in sequences]
    first = normalized[0]
    if any(sequence != first for sequence in normalized[1:]):
        return None
    return sequences[0]


def _normalize_sequence(sequence: list[Any]) -> list[str]:
    """将轴值序列统一转换为可比较的字符串列表。"""
    return [_stringify_cell_value(value) for value in sequence]


def _render_scatter_like_chart_table(
    x_sequences: list[list[Any]] | None,
    series_names: list[str],
    series_y_values: list[list[Any]],
    *,
    x_axis_title: str,
) -> str:
    """按共享或独立 X 轴布局，将散点图序列渲染为 HTML 表格。"""
    if not x_sequences or not series_names or len(x_sequences) != len(series_names):
        return ""

    shared_x_values = _get_shared_axis_values(x_sequences)
    if shared_x_values is not None:
        row_count = max(
            len(shared_x_values),
            max((len(values) for values in series_y_values), default=0),
        )
        if row_count == 0:
            return ""

        headers = [x_axis_title or ""] + series_names
        columns = [_stringify_series_values(shared_x_values)]
        columns.extend(_stringify_series_values(values) for values in series_y_values)
        return _render_html_table(headers, columns, row_count)

    headers = []
    columns = []
    row_count = 0
    for name, x_values, y_values in zip(series_names, x_sequences, series_y_values):
        headers.extend((f"{name} X", f"{name} Y"))
        columns.append(_stringify_series_values(x_values))
        columns.append(_stringify_series_values(y_values))
        row_count = max(row_count, len(x_values), len(y_values))

    if row_count == 0:
        return ""

    return _render_html_table(headers, columns, row_count)


def _render_bubble_chart_table(
    x_sequences: list[list[Any]] | None,
    series_names: list[str],
    series_y_values: list[list[Any]],
    series_sizes: list[list[Any]],
    *,
    x_axis_title: str,
) -> str:
    """按共享或独立 X 轴布局，将气泡图三维序列渲染为 HTML 表格。"""
    if (
        not x_sequences
        or not series_names
        or len(x_sequences) != len(series_names)
        or len(series_y_values) != len(series_names)
        or len(series_sizes) != len(series_names)
    ):
        return ""

    shared_x_values = _get_shared_axis_values(x_sequences)
    if shared_x_values is not None:
        row_count = max(
            len(shared_x_values),
            max((len(values) for values in series_y_values), default=0),
            max((len(values) for values in series_sizes), default=0),
        )
        if row_count == 0:
            return ""

        headers = [x_axis_title or ""]
        columns = [_stringify_series_values(shared_x_values)]
        for name, y_values, bubble_sizes in zip(series_names, series_y_values, series_sizes):
            headers.extend((name, f"{name} size"))
            columns.append(_stringify_series_values(y_values))
            columns.append(_stringify_series_values(bubble_sizes))
        return _render_html_table(headers, columns, row_count)

    headers = []
    columns = []
    row_count = 0
    for name, x_values, y_values, bubble_sizes in zip(
        series_names,
        x_sequences,
        series_y_values,
        series_sizes,
    ):
        headers.extend((f"{name} X", f"{name} Y", f"{name} size"))
        columns.append(_stringify_series_values(x_values))
        columns.append(_stringify_series_values(y_values))
        columns.append(_stringify_series_values(bubble_sizes))
        row_count = max(row_count, len(x_values), len(y_values), len(bubble_sizes))

    if row_count == 0:
        return ""

    return _render_html_table(headers, columns, row_count)


def _stringify_series_values(
    values: list[Any],
    *,
    date_hint: bool = False,
    date_1904: bool = False,
) -> list[str]:
    """将图表序列值批量转换为适合 HTML 输出的文本。"""
    return [
        _stringify_cell_value(
            value,
            date_hint=date_hint,
            date_1904=date_1904,
        )
        for value in values
    ]


def _stringify_cache_value(
    value: str | None,
    *,
    date_hint: bool = False,
    date_1904: bool = False,
) -> str:
    """规范化 OOXML 缓存值，并在需要时转换 Excel 日期序列号。"""
    if value in (None, ""):
        return ""

    if date_hint:
        try:
            serial = float(value)
        except (TypeError, ValueError):
            return value
        return _excel_serial_to_iso(serial, date_1904=date_1904) or value

    return value


def _stringify_cell_value(
    value: Any,
    *,
    date_hint: bool = False,
    date_1904: bool = False,
) -> str:
    """将工作簿单元格值转换为稳定文本，并保留日期时间语义。"""
    if value in (None, ""):
        return ""

    if isinstance(value, datetime):
        if date_hint and value.time() == time():
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()

    if date_hint and isinstance(value, (int, float)):
        return _excel_serial_to_iso(float(value), date_1904=date_1904) or _stringify_non_date_value(value)

    return _stringify_non_date_value(value)


def _excel_serial_to_iso(serial: float, *, date_1904: bool = False) -> str | None:
    """把 Excel 日期序列号转换为 ISO 文本，无效数值返回空结果。"""
    if not math.isfinite(serial):
        return None
    try:
        excel_value = from_excel(serial, MAC_EPOCH if date_1904 else WINDOWS_EPOCH)
    except (TypeError, ValueError, OverflowError):
        return None
    if isinstance(excel_value, datetime):
        if excel_value.time() == time():
            return excel_value.date().isoformat()
        return excel_value.isoformat(sep=" ")
    if isinstance(excel_value, date):
        return excel_value.isoformat()
    if isinstance(excel_value, time):
        return excel_value.isoformat()
    return str(excel_value)


def _stringify_non_date_value(value: Any) -> str:
    """将非日期值转换为文本，并去除整数浮点值末尾的小数部分。"""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _render_html_table(headers: list[str], columns: list[list[str]], row_count: int) -> str:
    """按表头、列数据和行数生成已转义的紧凑 HTML 表格。"""
    if row_count <= 0 or len(headers) != len(columns):
        return ""

    html_parts = ["<table><thead><tr>"]
    for header in headers:
        html_parts.append(f"<th>{escape(header)}</th>")
    html_parts.append("</tr></thead><tbody>")

    for row_idx in range(row_count):
        html_parts.append("<tr>")
        for column in columns:
            value = column[row_idx] if row_idx < len(column) else ""
            html_parts.append(f"<td>{escape(value)}</td>")
        html_parts.append("</tr>")

    html_parts.append("</tbody></table>")
    return "".join(html_parts)
