# Copyright (c) Opendatalab. All rights reserved.

"""把 Excel 97–2003 BIFF 工作簿转换为 MinerU 分页 model-list。"""

from __future__ import annotations

import collections
import re
from typing import Any, BinaryIO

from loguru import logger
from openpyxl import Workbook  # type: ignore[reportMissingModuleSource]
from openpyxl.cell.rich_text import CellRichText, TextBlock  # type: ignore[reportMissingModuleSource]
from openpyxl.cell.text import InlineFont  # type: ignore[reportMissingModuleSource]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[reportMissingModuleSource]

from ..errors import LegacyOfficeEncryptedError, LegacyOfficeResourceLimitError
from ..limits import MAX_GRID_SLOTS
from ..equation.mtef import decode_equation_native
from ..legacy.ole import BoundedOleReader
from ..spreadsheet.html import render_spreadsheet_table
from ..spreadsheet.models import AnchoredBlock, ExcelTable, FormulaMap, SheetImage
from ..spreadsheet.projector import SpreadsheetProjector
from ..streams import read_stream_bytes_from_start
from .....types import BlockType

from .models import XlsChart, XlsChartSheet, XlsRichText, XlsSheet, XlsWorkbook
from .parser import parse_xls_workbook

_XLS_EQUATION_STREAM_RE = re.compile(
    r"^(MBD[0-9A-F]{8})/Equation Native$",
    re.IGNORECASE,
)


def _read_embedded_equations(ole: BoundedOleReader) -> dict[str, str]:
    """读取 XLS embedding storages 中可安全解码的 Equation Native。"""

    equations: dict[str, str] = {}
    for stream_name in ole.stream_names(prefix="MBD"):
        match = _XLS_EQUATION_STREAM_RE.match(stream_name)
        if match is None:
            continue
        storage = match.group(1)
        latex = decode_equation_native(ole.read_stream(stream_name))
        if latex is None:
            logger.warning(
                "XLS_MTEF_FALLBACK: storage={!r} has an invalid or unsupported Equation Native stream",
                storage,
            )
            continue
        equations.setdefault(storage, latex)
    return equations


def _inline_font(rich_text: XlsRichText, start: int) -> InlineFont | None:
    """返回覆盖指定字符位置的 openpyxl 行内字体。"""

    for run in rich_text.runs:
        if run.start <= start < run.end:
            style = run.style
            return InlineFont(
                b=style.bold,
                i=style.italic,
                u="single" if style.underline else None,
                strike=style.strike,
                vertAlign=("superscript" if style.superscript else "subscript" if style.subscript else None),
            )
    return None


def _rich_text_boundaries(value: XlsRichText) -> list[int]:
    """收集富文本的起止边界并裁剪到有效字符范围。"""

    boundaries = {0, len(value.text)}
    for run in value.runs:
        boundaries.add(max(0, min(run.start, len(value.text))))
        boundaries.add(max(0, min(run.end, len(value.text))))
    return sorted(boundaries)


def _to_openpyxl_rich_text(value: XlsRichText) -> str | CellRichText:
    """把内部富文本转换为 XlsxConverter 已支持的 CellRichText。"""

    if not value.runs:
        return value.text
    parts: list[str | TextBlock] = []
    boundaries = _rich_text_boundaries(value)
    for start, end in zip(boundaries, boundaries[1:]):
        if start >= end:
            continue
        text = value.text[start:end]
        font = _inline_font(value, start)
        parts.append(TextBlock(font, text) if font is not None else text)
    return CellRichText(parts)


class _XlsPageBuilder(SpreadsheetProjector):
    """用轻量 openpyxl worksheet 适配器复用现有网格与 HTML 投影。"""

    def __init__(self, workbook_model: XlsWorkbook) -> None:
        """初始化解析结果映射和全工作簿网格预算。"""

        super().__init__(include_hidden_sheets=False)
        self.workbook_model = workbook_model
        self._sheet_by_title: dict[str, XlsSheet] = {}
        self._active_xls_sheet: XlsSheet | None = None
        self._used_cells: set[tuple[int, int]] = set()
        self._grid_slots = 0

    def _build_openpyxl_workbook(self) -> Workbook:
        """把 BIFF 语义模型投影成不落盘的 worksheet 对象。"""

        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)
        for sheet_model in self.workbook_model.sheets:
            worksheet = workbook.create_sheet(sheet_model.name)
            self._sheet_by_title[worksheet.title] = sheet_model
            if not sheet_model.visible:
                worksheet.sheet_state = Worksheet.SHEETSTATE_HIDDEN
            for cell_model in sheet_model.cells.values():
                cell = worksheet.cell(row=cell_model.row + 1, column=cell_model.col + 1)
                cell.value = _to_openpyxl_rich_text(cell_model.value)
                if cell_model.hyperlink:
                    cell.hyperlink = cell_model.hyperlink
            for row_first, col_first, row_last, col_last in sheet_model.merges:
                worksheet.merge_cells(
                    start_row=row_first + 1,
                    start_column=col_first + 1,
                    end_row=row_last + 1,
                    end_column=col_last + 1,
                )
        return workbook

    def _ensure_workbook(self) -> Workbook:
        """惰性构造一次 openpyxl 适配工作簿。"""

        workbook = getattr(self, "workbook", None)
        if workbook is None:
            workbook = self._build_openpyxl_workbook()
            self.workbook = workbook
        return workbook

    def render_chart_selection(
        self,
        sheet_name: str,
        rows: tuple[int, ...] | list[int],
        cols: tuple[int, ...] | list[int],
    ) -> str | None:
        """把指定 worksheet 行列选择渲染为共享 chart HTML。"""

        if not rows or not cols:
            return None
        grid_slots = len(rows) * len(cols)
        self._grid_slots += grid_slots
        if self._grid_slots > MAX_GRID_SLOTS:
            raise LegacyOfficeResourceLimitError(f"workbook extent exceeds max_grid_slots={MAX_GRID_SLOTS}")
        workbook = self._ensure_workbook()
        if sheet_name not in workbook.sheetnames:
            return None
        sheet = workbook[sheet_name]
        table = self._build_synthetic_table_from_sheet_selection(
            sheet,
            list(rows),
            list(cols),
        )
        return render_spreadsheet_table(table)

    def _chart_sheet_page(self, chart_sheet: XlsChartSheet) -> list[dict[str, Any]]:
        """把独立 chart sheet 投影为仅含 chart block 的逻辑页。"""

        if chart_sheet.source_sheet_name is None:
            return []
        content = self.render_chart_selection(
            chart_sheet.source_sheet_name,
            chart_sheet.source_rows,
            chart_sheet.source_cols,
        )
        return [{"type": BlockType.CHART, "content": content}] if content else []

    def build_pages(self) -> list[list[dict[str, Any]]]:
        """按原目录顺序生成可见 worksheet/chart sheet 页面。"""

        self.workbook = self._build_openpyxl_workbook()
        pages_by_name: dict[str, list[dict[str, Any]]] = {}
        for worksheet in self._iter_sheets_to_convert():
            self._active_xls_sheet = self._sheet_by_title.get(worksheet.title)
            self.cur_page = []
            self._convert_sheet(worksheet)
            pages_by_name[worksheet.title] = self.cur_page

        ordered_pages: list[tuple[int, int, str, list[dict[str, Any]]]] = []
        for index, sheet in enumerate(self.workbook_model.sheets):
            if not sheet.visible:
                continue
            order = sheet.order if sheet.order >= 0 else index
            ordered_pages.append((order, 0, sheet.name, pages_by_name.get(sheet.name, [])))
        for index, chart_sheet in enumerate(self.workbook_model.chart_sheets):
            if not chart_sheet.visible:
                continue
            order = chart_sheet.order if chart_sheet.order >= 0 else len(ordered_pages) + index
            ordered_pages.append((order, 1, chart_sheet.name, self._chart_sheet_page(chart_sheet)))
        ordered_pages.sort(key=lambda item: (item[0], item[1]))
        sheet_pages = [(name, page) for _order, _kind, name, page in ordered_pages]
        if self._should_emit_sheet_titles([page for _, page in sheet_pages]):
            self._prepend_sheet_titles(sheet_pages)
        return [page for _, page in sheet_pages]

    def _collect_sheet_images(self, sheet: Worksheet) -> list[SheetImage]:
        """返回解析器已经绑定到当前 sheet 的图片。"""

        sheet_model = self._sheet_by_title.get(sheet.title)
        if sheet_model is None:
            return []
        return [
            SheetImage(
                anchor=(image.row, image.col),
                image_base64=image.image_base64,
            )
            for image in sheet_model.images
        ]

    def _map_math_formulas_to_cells(self, sheet: Worksheet) -> FormulaMap:
        """把 legacy Equation Editor 公式映射到表格 cell anchor。"""

        math_map: dict[tuple[int, int], list[str]] = collections.defaultdict(list)
        sheet_model = self._sheet_by_title.get(sheet.title)
        if sheet_model is None:
            return math_map
        for equation in sheet_model.equations:
            math_map[(equation.row, equation.col)].append(equation.latex)
        return math_map

    def _find_tables_in_sheet(
        self,
        sheet: Worksheet,
    ) -> tuple[set[tuple[int, int]], list[tuple[tuple[int, int], int, dict]]]:
        """记录表格已吸收坐标，供独立公式去重。"""

        used_cells, artifacts = super()._find_tables_in_sheet(sheet)
        self._used_cells = used_cells
        return used_cells, artifacts

    def _find_data_tables(self, sheet: Worksheet) -> list[ExcelTable]:
        """复用 XLSX 区域发现并累计限制实际物化网格。"""

        tables = super()._find_data_tables(sheet)
        self._grid_slots += sum(table.num_rows * table.num_cols for table in tables)
        if self._grid_slots > MAX_GRID_SLOTS:
            raise LegacyOfficeResourceLimitError(f"workbook extent exceeds max_grid_slots={MAX_GRID_SLOTS}")
        return tables

    def _chart_block(self, sheet: Worksheet, chart: XlsChart) -> dict[str, Any] | None:
        """把简单 chart 引用转换成数据表，并保留可用预览图片。"""

        if chart.source_rows and chart.source_cols:
            content = self.render_chart_selection(
                sheet.title,
                chart.source_rows,
                chart.source_cols,
            )
            if content is None:
                return None
            block = {
                "type": BlockType.CHART,
                "content": content,
            }
            if chart.image_base64:
                block["image_base64"] = chart.image_base64
            return block
        if chart.image_base64:
            return {
                "type": BlockType.CHART,
                "content": "",
                "image_base64": chart.image_base64,
            }
        return None

    def _find_charts_in_sheet(
        self,
        sheet: Worksheet,
    ) -> list[AnchoredBlock]:
        """按 cell anchor 输出当前工作表的 legacy chart blocks。"""

        sheet_model = self._sheet_by_title.get(sheet.title)
        if sheet_model is None:
            return []
        artifacts = []
        for order, chart in enumerate(sheet_model.charts):
            block = self._chart_block(sheet, chart)
            if block is None:
                continue
            artifacts.append(((chart.row, chart.col), 10_000 + order, block))
        for order, equation in enumerate(sheet_model.equations):
            anchor = (equation.row, equation.col)
            if anchor in self._used_cells:
                continue
            artifacts.append(
                (
                    anchor,
                    20_000 + order,
                    {"type": BlockType.EQUATION, "content": equation.latex},
                )
            )
        return artifacts


def render_xls_chart_html(
    workbook: XlsWorkbook,
    sheet_name: str,
    rows: tuple[int, ...] | list[int],
    cols: tuple[int, ...] | list[int],
) -> str | None:
    """为 DOC/PPT 嵌入式 chart 复用 XLS 的稳定 HTML 投影。"""

    return _XlsPageBuilder(workbook).render_chart_selection(
        sheet_name,
        rows,
        cols,
    )


class XlsConverter:
    """将 Excel 97–2003 OLE/BIFF 二进制流转换为 model-list。"""

    def __init__(self) -> None:
        """初始化空分页输出。"""

        self.pages: list[list[dict[str, Any]]] = []

    def convert(self, file_binary: BinaryIO) -> None:
        """读取 Workbook/Book stream，解析 BIFF 并生成可见 sheet 页面。"""

        file_bytes = read_stream_bytes_from_start(file_binary)
        with BoundedOleReader(file_bytes) as ole:
            if ole.has_stream("EncryptionInfo") or ole.has_stream("EncryptedPackage"):
                raise LegacyOfficeEncryptedError("password-protected XLS is not supported")
            if ole.has_stream("Workbook"):
                workbook_stream = ole.read_stream("Workbook")
            else:
                workbook_stream = ole.read_stream("Book")
            native_equations = _read_embedded_equations(ole)
            workbook = parse_xls_workbook(
                workbook_stream,
                native_equations=native_equations,
            )
        builder = _XlsPageBuilder(workbook)
        self.pages = builder.build_pages()
        logger.debug("XLS parsing produced {} visible sheet pages", len(self.pages))
