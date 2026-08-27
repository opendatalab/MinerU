# Copyright (c) Opendatalab. All rights reserved.
"""XLS 与 XLSX 复用的中立工作表投影器。"""

from __future__ import annotations

import collections
import html
from collections.abc import Iterator
from typing import Any

from loguru import logger
from openpyxl.cell.rich_text import CellRichText
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .....types import BlockType
from ..._shared.hyperlink import OFFICE_EXTERNAL_HYPERLINK_SCHEMES, sanitize_hyperlink_target
from .html import EQUATION_BOOKENDS, render_spreadsheet_table
from .models import AnchoredBlock, DataRegion, ExcelCell, ExcelTable, FormulaMap, SheetImage

AUTO_GAP_TOLERANCE_CANDIDATES = (0, 1, 2)
AUTO_GAP_TOLERANCE_PREFERENCE = {1: 0, 0: 1, 2: 2}
AUTO_GAP_TOLERANCE_PREFERENCE_MARGIN = 0.15


class _MergedCellLookup:
    """按行缓存合并单元格范围，避免解析时反复扫描 openpyxl 合并区域。"""

    def __init__(self, sheet: Worksheet):
        """从工作表合并区域构建 0-based 坐标索引。"""
        self._merged_row_intervals: dict[int, list[tuple[int, int]]] = collections.defaultdict(list)
        self._hidden_row_intervals: dict[int, list[tuple[int, int]]] = collections.defaultdict(list)
        self._anchor_spans: dict[tuple[int, int], tuple[int, int]] = {}

        for merged in sheet.merged_cells.ranges:
            min_row = merged.min_row - 1
            max_row = merged.max_row - 1
            min_col = merged.min_col - 1
            max_col = merged.max_col - 1

            self._anchor_spans[(min_row, min_col)] = (
                max_row - min_row + 1,
                max_col - min_col + 1,
            )

            for row in range(min_row, max_row + 1):
                self._merged_row_intervals[row].append((min_col, max_col))
                hidden_start_col = min_col + 1 if row == min_row else min_col
                if hidden_start_col <= max_col:
                    self._hidden_row_intervals[row].append((hidden_start_col, max_col))

        for intervals in self._merged_row_intervals.values():
            intervals.sort()
        for intervals in self._hidden_row_intervals.values():
            intervals.sort()

    @staticmethod
    def _contains_interval(
        row_intervals: dict[int, list[tuple[int, int]]],
        row: int,
        col: int,
    ) -> bool:
        """判断 0-based 坐标是否落入指定行的任一列区间。"""
        for start_col, end_col in row_intervals.get(row, []):
            if start_col <= col <= end_col:
                return True
            if start_col > col:
                break
        return False

    def contains_merged_cell(self, row: int, col: int) -> bool:
        """判断 0-based 坐标是否属于任一合并区域。"""
        return self._contains_interval(self._merged_row_intervals, row, col)

    def is_hidden_merged_cell(self, row: int, col: int) -> bool:
        """判断 0-based 坐标是否为合并区域内非左上角的隐藏格。"""
        return self._contains_interval(self._hidden_row_intervals, row, col)

    def get_anchor_span(self, row: int, col: int) -> tuple[int, int]:
        """返回合并区域左上角坐标对应的 rowspan/colspan，非合并锚点返回 1x1。"""
        return self._anchor_spans.get((row, col), (1, 1))


class SpreadsheetProjector:
    """把 openpyxl 工作表投影为稳定的分页 model-list。"""

    def __init__(
        self,
        *,
        treat_singleton_as_text: bool = True,
        gap_tolerance: int | None = None,
        include_hidden_sheets: bool = False,
    ) -> None:
        """保存工作表投影配置并初始化无格式专属依赖的运行状态。"""
        self.treat_singleton_as_text = treat_singleton_as_text
        self.gap_tolerance = gap_tolerance
        self.include_hidden_sheets = include_hidden_sheets
        self._reset_projection_state()

    def _reset_projection_state(self) -> None:
        """重置工作簿、分页和逐 sheet 的共享投影状态。"""
        self.workbook: Workbook | None = None
        self.pages: list[list[dict[str, Any]]] = []
        self.cur_page: list[dict[str, Any]] = []
        self.math_map: FormulaMap = {}
        self.sheet_images: list[SheetImage] = []
        self.table_image_map: dict[tuple[int, int], list[str]] = collections.defaultdict(list)
        self._merged_cell_lookup_cache: dict[int, _MergedCellLookup] = {}

    def _prepare_sheet_assets(self, sheet: Worksheet) -> None:
        """准备普通公式和图片，并构建表格单元格使用的媒体映射。"""
        self.math_map = self._map_math_formulas_to_cells(sheet)
        self.sheet_images = self._collect_sheet_images(sheet)
        self.table_image_map = collections.defaultdict(list)
        for image in self.sheet_images:
            row, col = image.anchor
            if row is None or col is None:
                continue
            if image.latex:
                self.table_image_map[(row, col)].append(EQUATION_BOOKENDS.format(EQ=image.latex))
            elif image.image_base64:
                self.table_image_map[(row, col)].append(f'<img src="{image.image_base64}" />')

    def _convert_sheet(self, sheet: Worksheet) -> None:
        """按表格、图表、附加素材和独立图片的稳定顺序投影一个工作表。"""
        self._prepare_sheet_assets(sheet)
        used_cells, visual_artifacts = self._find_tables_in_sheet(sheet)
        visual_artifacts.extend(self._find_charts_in_sheet(sheet))
        visual_artifacts.extend(self._find_additional_visual_artifacts(used_cells))
        for _, _, block in sorted(
            visual_artifacts,
            key=lambda item: (item[0][0], item[0][1], item[1]),
        ):
            self.cur_page.append(block)
        self._find_images_in_sheet(used_cells)

    def _map_math_formulas_to_cells(self, sheet: Worksheet) -> FormulaMap:
        """返回当前工作表按 0-based cell anchor 分组的公式。"""
        return {}

    def _collect_sheet_images(self, sheet: Worksheet) -> list[SheetImage]:
        """返回当前工作表按 anchor 排序的图片或图片公式。"""
        return []

    def _find_charts_in_sheet(self, sheet: Worksheet) -> list[AnchoredBlock]:
        """返回当前工作表的格式专属图表 blocks。"""
        return []

    def _find_additional_visual_artifacts(
        self,
        used_cells: set[tuple[int, int]],
    ) -> list[AnchoredBlock]:
        """返回未被表格吸收的格式专属公式或图片 blocks。"""
        return []

    def _resolve_cell_image(self, raw_cell_text: str) -> str:
        """解析格式专属的单元格图片函数，默认不产生媒体。"""
        return ""

    def _iter_sheets_to_convert(self) -> Iterator[Worksheet]:
        """按工作簿顺序遍历允许输出的可见工作表。"""
        if self.workbook is None:
            return

        for sheet in self.workbook.worksheets:
            if not self.include_hidden_sheets and sheet.sheet_state != Worksheet.SHEETSTATE_VISIBLE:
                logger.debug(f"跳过隐藏工作表：{sheet.title}")
                continue
            yield sheet

    @staticmethod
    def _build_sheet_title_block(sheet_title: str) -> dict:
        """构造工作表标题块，复用 Office 标题渲染链路输出 Markdown 标题。"""
        return {
            "type": BlockType.PARAGRAPH_TITLE,
            "level": 2,
            "content": sheet_title,
        }

    @staticmethod
    def _should_emit_sheet_titles(pages: list[list[dict]]) -> bool:
        """仅当存在多个非空输出 sheet 时才添加标题，避免单表或空表噪声。"""
        return sum(1 for page in pages if page) > 1

    def _prepend_sheet_titles(self, sheet_pages: list[tuple[str, list[dict]]]) -> None:
        """将 sheet 标题插入每个非空 page 开头，不参与表格/图表视觉排序。"""
        for sheet_title, page in sheet_pages:
            if not page:
                continue
            page.insert(0, self._build_sheet_title_block(sheet_title))

    def _get_block_sort_anchor(self, row: int | None, col: int | None) -> tuple[int, int]:
        """把缺失 anchor 稳定放到全部有效工作表坐标之后。"""
        if row is None or col is None:
            return (10**9, 10**9)
        return row, col

    def _build_block_from_excel_table(self, excel_table: ExcelTable) -> dict:
        """按 singleton 规则把表格 IR 投影为文本或表格 block。"""
        if self.treat_singleton_as_text and len(excel_table.data) == 1 and self._can_render_singleton_as_text(excel_table):
            return {
                "type": BlockType.TEXT,
                "content": excel_table.data[0].text,
            }

        return {
            "type": BlockType.TABLE,
            "content": render_spreadsheet_table(excel_table),
        }

    def _find_tables_in_sheet(self, sheet: Worksheet) -> tuple[set[tuple[int, int]], list[tuple[tuple[int, int], int, dict]]]:
        """发现当前 sheet 表格并返回已吸收 cell 与锚定 blocks。"""
        used_cells = set()
        visual_artifacts = []
        if self.workbook is not None:
            tables = self._find_data_tables(sheet)  # 检测工作表中的所有数据表格

            for order, excel_table in enumerate(tables):
                # Record used cells
                anchor_c, anchor_r = excel_table.anchor
                for cell in excel_table.data:
                    source_row, source_col = self._resolve_excel_cell_source_position(
                        excel_table.anchor,
                        cell,
                    )
                    used_cells.add((source_row, source_col))

                visual_artifacts.append(
                    (
                        self._get_block_sort_anchor(anchor_r, anchor_c),
                        order,
                        self._build_block_from_excel_table(excel_table),
                    )
                )

        return used_cells, visual_artifacts

    def _build_excel_cell(
        self,
        sheet: Worksheet,
        display_row: int,
        display_col: int,
        source_row: int,
        source_col: int,
        row_span: int = 1,
        col_span: int = 1,
    ) -> ExcelCell:
        """把源工作表单元格完整物化为中立 ExcelCell。"""
        cell = sheet.cell(row=source_row + 1, column=source_col + 1)
        raw_cell_text = str(cell.value) if cell.value is not None else ""
        cell_text = ""
        text_is_html = False
        media_content = []
        if "DISPIMG" in raw_cell_text:
            cell_image = self._resolve_cell_image(raw_cell_text)
            if cell_image:
                media_content.append(cell_image)
        else:
            cell_text, text_is_html = self._cell_value_to_html(cell)
        media_content.extend(self.table_image_map.get((source_row, source_col), []))

        return ExcelCell(
            row=display_row,
            col=display_col,
            text=cell_text,
            row_span=row_span,
            col_span=col_span,
            styles=self._extract_cell_style(cell),
            media=media_content,
            equations=list(self.math_map.get((source_row, source_col), [])),
            text_is_html=text_is_html,
            source_row=source_row,
            source_col=source_col,
        )

    def _build_synthetic_table_from_sheet_selection(self, sheet: Worksheet, rows: list[int], cols: list[int]) -> ExcelTable:
        """把指定源行列选择物化为紧凑的表格 IR。"""
        selected_coords = {(row, col) for row in rows for col in cols}
        hidden_merge_cells = set()
        merge_spans = {}

        for mr in sheet.merged_cells.ranges:
            top_left = (mr.min_row - 1, mr.min_col - 1)
            if top_left not in selected_coords:
                continue

            selected_rows = [row for row in rows if mr.min_row - 1 <= row <= mr.max_row - 1]
            selected_cols = [col for col in cols if mr.min_col - 1 <= col <= mr.max_col - 1]
            if not selected_rows or not selected_cols:
                continue

            merge_spans[top_left] = (len(selected_rows), len(selected_cols))
            for row in selected_rows:
                for col in selected_cols:
                    if (row, col) != top_left:
                        hidden_merge_cells.add((row, col))

        data = []
        for display_row, source_row in enumerate(rows):
            for display_col, source_col in enumerate(cols):
                if (source_row, source_col) in hidden_merge_cells:
                    continue

                row_span, col_span = merge_spans.get((source_row, source_col), (1, 1))
                data.append(
                    self._build_excel_cell(
                        sheet,
                        display_row,
                        display_col,
                        source_row,
                        source_col,
                        row_span=row_span,
                        col_span=col_span,
                    )
                )

        return ExcelTable(
            anchor=(cols[0], rows[0]),
            num_rows=len(rows),
            num_cols=len(cols),
            data=data,
        )

    def _resolve_excel_cell_source_position(
        self,
        table_anchor: tuple[int, int],
        excel_cell: ExcelCell | None,
        row: int | None = None,
        col: int | None = None,
    ) -> tuple[int, int]:
        """优先使用显式源坐标，否则通过表格 anchor 还原源坐标。"""
        if excel_cell is not None:
            if excel_cell.source_row is not None and excel_cell.source_col is not None:
                return excel_cell.source_row, excel_cell.source_col
            row = excel_cell.row
            col = excel_cell.col

        if row is None or col is None:
            raise ValueError("row and col must be provided when excel_cell is None")

        return table_anchor[1] + row, table_anchor[0] + col

    def _can_render_singleton_as_text(self, excel_table: ExcelTable) -> bool:
        """判断单格表是否可安全降级为普通文本 block。"""
        cell = excel_table.data[0]
        return (
            cell.row_span == 1
            and cell.col_span == 1
            and not cell.media
            and not cell.text_is_html
            and not cell.equations
        )

    def _cell_has_semantic_content(self, excel_table: ExcelTable, cell: ExcelCell) -> bool:
        """判断单元格是否包含文本、媒体或公式语义。"""
        return bool(
            cell.text.strip()
            or any(media.strip() for media in cell.media)
            or cell.equations
        )

    def _get_table_semantic_positions(self, excel_table: ExcelTable) -> set[tuple[int, int]]:
        """返回表格内具有语义内容的源工作表坐标。"""
        semantic_positions = set()
        for cell in excel_table.data:
            if not self._cell_has_semantic_content(excel_table, cell):
                continue
            semantic_positions.add(
                self._resolve_excel_cell_source_position(
                    excel_table.anchor,
                    excel_cell=cell,
                )
            )
        return semantic_positions

    def _filter_semantic_subset_tables(self, tables: list[ExcelTable]) -> list[ExcelTable]:
        """删除语义坐标严格包含于其它候选的重复表格。"""
        semantic_positions = [self._get_table_semantic_positions(table) for table in tables]
        filtered_tables = []

        for table_idx, table in enumerate(tables):
            if any(
                semantic_positions[table_idx] < semantic_positions[other_idx]
                for other_idx in range(len(tables))
                if other_idx != table_idx
            ):
                continue
            filtered_tables.append(table)

        return filtered_tables

    def _build_table_content_mask(self, excel_table: ExcelTable) -> list[list[bool]]:
        """构造包含合并跨度的表格语义内容掩码。"""
        mask = [[False for _ in range(excel_table.num_cols)] for _ in range(excel_table.num_rows)]
        for cell in excel_table.data:
            if not self._cell_has_semantic_content(excel_table, cell):
                continue
            for row_idx in range(cell.row, min(cell.row + cell.row_span, excel_table.num_rows)):
                for col_idx in range(cell.col, min(cell.col + cell.col_span, excel_table.num_cols)):
                    mask[row_idx][col_idx] = True
        return mask

    @staticmethod
    def _count_max_consecutive_true(flags: list[bool]) -> int:
        """返回布尔序列中最长连续真值长度。"""
        max_count = 0
        current = 0
        for flag in flags:
            if flag:
                current += 1
                max_count = max(max_count, current)
            else:
                current = 0
        return max_count

    @staticmethod
    def _is_real_singleton_table(excel_table: ExcelTable) -> bool:
        """判断候选是否是单格且不可进一步拆分的真实表格。"""
        if excel_table.num_rows != 1 or excel_table.num_cols != 1 or len(excel_table.data) != 1:
            return False
        cell = excel_table.data[0]
        return cell.row_span == 1 and cell.col_span == 1

    def _summarize_table_for_gap_selection(self, excel_table: ExcelTable) -> dict[str, float | int | bool]:
        """计算 gap 候选评分使用的单表形态指标。"""
        table_area = excel_table.num_rows * excel_table.num_cols
        content_mask = self._build_table_content_mask(excel_table)
        content_area = sum(sum(1 for flag in row if flag) for row in content_mask)
        blank_ratio = 1.0 - (content_area / max(table_area, 1))

        interior_blank_rows = [not any(content_mask[row_idx]) for row_idx in range(1, max(excel_table.num_rows - 1, 1))]
        interior_blank_cols = [
            not any(content_mask[row_idx][col_idx] for row_idx in range(excel_table.num_rows))
            for col_idx in range(1, max(excel_table.num_cols - 1, 1))
        ]
        if excel_table.num_rows <= 2:
            interior_blank_rows = []
        if excel_table.num_cols <= 2:
            interior_blank_cols = []

        interior_blank_row_count = sum(interior_blank_rows)
        interior_blank_col_count = sum(interior_blank_cols)
        max_consecutive_interior_blank_lines = max(
            self._count_max_consecutive_true(interior_blank_rows),
            self._count_max_consecutive_true(interior_blank_cols),
        )

        return {
            "table_area": table_area,
            "content_area": content_area,
            "blank_ratio": blank_ratio,
            "interior_blank_row_count": interior_blank_row_count,
            "interior_blank_col_count": interior_blank_col_count,
            "max_consecutive_interior_blank_lines": max_consecutive_interior_blank_lines,
            "real_singleton": self._is_real_singleton_table(excel_table),
        }

    def _summarize_candidate_tables(self, tables: list[ExcelTable]) -> dict[str, float | int]:
        """汇总一组 gap 候选表格的惩罚指标。"""
        table_count = len(tables)
        real_singleton_count = 0
        severe_separator_count = 0
        sparse_large_table_count = 0
        total_area = 0
        weighted_blank_numerator = 0.0
        total_interior_blank_lines = 0
        total_possible_interior_lines = 0
        row_cover_count = collections.Counter()

        for table in tables:
            table_summary = self._summarize_table_for_gap_selection(table)
            table_area = int(table_summary["table_area"])
            blank_ratio = float(table_summary["blank_ratio"])
            interior_blank_row_count = int(table_summary["interior_blank_row_count"])
            interior_blank_col_count = int(table_summary["interior_blank_col_count"])
            max_consecutive_interior_blank_lines = int(table_summary["max_consecutive_interior_blank_lines"])

            total_area += table_area
            weighted_blank_numerator += table_area * blank_ratio
            total_interior_blank_lines += interior_blank_row_count + interior_blank_col_count
            total_possible_interior_lines += max(table.num_rows - 2, 0) + max(table.num_cols - 2, 0)
            for row_idx in range(table.anchor[1], table.anchor[1] + table.num_rows):
                row_cover_count[row_idx] += 1

            if bool(table_summary["real_singleton"]):
                real_singleton_count += 1
            if table_area >= 6 and blank_ratio > 0.35:
                sparse_large_table_count += 1
            if max_consecutive_interior_blank_lines >= 2:
                severe_separator_count += 1

        occupied_row_count = max(len(row_cover_count), 1)
        row_overlap_excess_ratio = sum(max(0, count - 1) for count in row_cover_count.values()) / occupied_row_count

        return {
            "real_singleton_ratio": real_singleton_count / max(table_count, 1),
            "weighted_blank_ratio": weighted_blank_numerator / max(total_area, 1),
            "interior_blank_line_ratio": total_interior_blank_lines / max(total_possible_interior_lines, 1),
            "sparse_large_table_ratio": sparse_large_table_count / max(table_count, 1),
            "severe_separator_count": severe_separator_count,
            "row_overlap_excess_ratio": row_overlap_excess_ratio,
        }

    def _select_best_gap_candidate(self, sheet: Worksheet) -> tuple[int, float, list[ExcelTable]]:
        """按固定候选与偏好顺序选择最稳定的 gap tolerance。"""
        candidates = []
        for gap_tolerance in AUTO_GAP_TOLERANCE_CANDIDATES:
            raw_tables = self._find_data_tables_with_gap_raw(sheet, gap_tolerance)
            summary = self._summarize_candidate_tables(raw_tables)
            penalty = (
                6.0 * int(summary["severe_separator_count"])
                + 2.5 * float(summary["interior_blank_line_ratio"])
                + 1.5 * float(summary["sparse_large_table_ratio"])
                + 1.0 * float(summary["real_singleton_ratio"])
                + 0.5 * float(summary["weighted_blank_ratio"])
                + 1.0 * float(summary["row_overlap_excess_ratio"])
            )
            candidates.append(
                {
                    "gap_tolerance": gap_tolerance,
                    "penalty": penalty,
                    "tables": self._filter_semantic_subset_tables(raw_tables),
                    **summary,
                }
            )

        min_penalty = min(float(candidate["penalty"]) for candidate in candidates)
        near_best_candidates = [
            candidate
            for candidate in candidates
            if float(candidate["penalty"]) <= (min_penalty + AUTO_GAP_TOLERANCE_PREFERENCE_MARGIN)
        ]

        best_candidate = min(
            near_best_candidates,
            key=lambda candidate: (
                int(candidate["severe_separator_count"]),
                AUTO_GAP_TOLERANCE_PREFERENCE[int(candidate["gap_tolerance"])],
                float(candidate["interior_blank_line_ratio"]),
                float(candidate["penalty"]),
            ),
        )
        return (
            int(best_candidate["gap_tolerance"]),
            float(best_candidate["penalty"]),
            best_candidate["tables"],
        )

    def _select_best_tables(self, sheet: Worksheet) -> list[ExcelTable]:
        """选择并记录当前工作表的最佳表格候选集合。"""
        gap_tolerance, penalty, tables = self._select_best_gap_candidate(sheet)
        logger.debug(
            "Selected gap_tolerance={} for sheet '{}' with penalty={:.4f}",
            gap_tolerance,
            sheet.title,
            penalty,
        )
        return tables

    def _find_images_in_sheet(self, used_cells: set[tuple[int, int]] | None = None) -> None:
        """输出没有被表格吸收且不是公式载体的独立图片。"""
        if self.workbook is not None:
            for image in self.sheet_images:
                r, c = image.anchor
                if used_cells and r is not None and c is not None and (r, c) in used_cells:
                    continue

                if image.latex:
                    continue
                if image.image_base64:
                    self.cur_page.append(
                        {
                            "type": BlockType.IMAGE,
                            "image_base64": image.image_base64,
                        }
                    )

    def _find_data_tables(self, sheet: Worksheet) -> list[ExcelTable]:
        """在 Excel 工作表中查找所有紧凑的矩形数据表格。

        参数：
            sheet: 待解析的 Excel 工作表。

        返回：
            表示所有数据表格的 ExcelTable 对象列表。
        """
        if self.gap_tolerance is None:
            return self._select_best_tables(sheet)
        return self._find_data_tables_with_gap(sheet, self.gap_tolerance)

    def _find_data_tables_with_gap(self, sheet: Worksheet, gap_tolerance: int) -> list[ExcelTable]:
        """按固定 gap 发现表格并移除语义子集候选。"""
        return self._filter_semantic_subset_tables(self._find_data_tables_with_gap_raw(sheet, gap_tolerance))

    def _find_data_tables_with_gap_raw(self, sheet: Worksheet, gap_tolerance: int) -> list[ExcelTable]:
        """在固定 gap_tolerance 下查找工作表中的所有数据表格。"""
        bounds: DataRegion = self._find_true_data_bounds(sheet)  # 获取真实数据边界
        tables: list[ExcelTable] = []  # 存储已发现的表格
        visited: set[tuple[int, int]] = set()  # 记录已访问的单元格

        # 仅遍历已存在且有值的单元格，避免 iter_rows 在稀疏大表上创建大量空单元格。
        for ri, rj in self._get_non_empty_cell_positions(sheet, bounds):
            # 跳过已访问的单元格
            if (ri, rj) in visited:
                continue

            # 从当前单元格出发，通过洪水填充算法确定所属表格的边界
            table_bounds, visited_cells = self._find_table_bounds(
                sheet,
                ri,
                rj,
                bounds.max_row,
                bounds.max_col,
                gap_tolerance,
            )
            visited.update(visited_cells)  # 将已访问单元格加入全局记录
            tables.append(table_bounds)

        return tables

    def _get_non_empty_cell_positions(
        self,
        sheet: Worksheet,
        bounds: DataRegion,
    ) -> list[tuple[int, int]]:
        """按行列顺序返回真实边界内已有值单元格的 0-based 坐标。"""
        positions = []
        for cell in sheet._cells.values():
            if cell.value is None:
                continue
            if not (bounds.min_row <= cell.row <= bounds.max_row and bounds.min_col <= cell.column <= bounds.max_col):
                continue
            positions.append((cell.row - 1, cell.column - 1))
        return sorted(positions)

    def _find_true_data_bounds(self, sheet: Worksheet) -> DataRegion:
        """查找工作表中真实的数据边界（最小/最大行列）。

        该函数扫描所有单元格，找到包含所有非空单元格或合并单元格区域的
        最小矩形范围，返回边界的行列索引。

        参数：
            sheet: 待分析的工作表。

        返回：
            覆盖所有数据和合并单元格的最小矩形区域 DataRegion。
            若工作表为空，则默认返回 (1, 1, 1, 1)。
        """
        min_row, min_col = None, None
        max_row, max_col = 0, 0

        # 遍历所有有值的单元格，动态更新边界
        for cell in sheet._cells.values():
            if cell.value is not None:
                r, c = cell.row, cell.column
                min_row = r if min_row is None else min(min_row, r)
                min_col = c if min_col is None else min(min_col, c)
                max_row = max(max_row, r)
                max_col = max(max_col, c)

        # 将合并单元格的范围也纳入边界计算
        for merged in sheet.merged_cells.ranges:
            min_row = merged.min_row if min_row is None else min(min_row, merged.min_row)
            min_col = merged.min_col if min_col is None else min(min_col, merged.min_col)
            max_row = max(max_row, merged.max_row)
            max_col = max(max_col, merged.max_col)

        # 若工作表中没有任何数据，默认返回 (1, 1, 1, 1)
        if min_row is None or min_col is None:
            min_row = min_col = max_row = max_col = 1

        return DataRegion(min_row, max_row, min_col, max_col)

    def _find_table_bounds(
        self,
        sheet: Worksheet,
        start_row: int,
        start_col: int,
        max_row: int,
        max_col: int,
        gap_tolerance: int,
    ) -> tuple[ExcelTable, set[tuple[int, int]]]:
        """使用洪水填充（BFS）策略确定表格边界。

        该方法通过广度优先搜索（BFS）算法识别 Excel 工作表中连续的非空单元格区域，
        能够准确检测非矩形表格（如 L 形、错位列等），并支持通过间隔容忍度
        连接相邻但不直接相连的单元格。

        算法分两个阶段执行：
        1. 洪水填充阶段：使用 BFS 从给定位置出发，找出所有相连的单元格。
        2. 数据提取阶段：构建矩形边界框并提取单元格数据，正确处理合并单元格。

        参数：
            sheet: 待分析的 Excel 工作表。
            start_row: 洪水填充起始行索引（从0开始）。
            start_col: 洪水填充起始列索引（从0开始）。
            max_row: 工作表中可考虑的最大行索引（从0开始）。
            max_col: 工作表中可考虑的最大列索引（从0开始）。
            gap_tolerance: 允许跨越空白单元格查找邻居的最大间隔。

        返回：
            一个元组，包含：
                - ExcelTable：表示检测到的表格对象，含锚点位置、尺寸和单元格数据。
                - set[tuple[int, int]]：洪水填充期间访问的所有 (行, 列) 元组集合，
                  用于防止重复扫描。

        说明：
            该方法遵循 GAP_TOLERANCE 选项，允许在容忍距离内将被空单元格隔开的
            单元格视为同一表格的一部分。
        """

        # BFS 队列，存储待处理的 (行, 列) 坐标
        queue = collections.deque([(start_row, start_col)])

        # 记录当前表格内已访问的单元格（避免重复加入队列）
        # 调用方维护全局 visited 集合，防止重复启动新表格
        table_cells: set[tuple[int, int]] = set()
        table_cells.add((start_row, start_col))

        # 动态记录当前表格的行列边界
        min_r, max_r = start_row, start_row
        min_c, max_c = start_col, start_col
        merged_lookup = self._get_merged_cell_lookup(sheet)

        def has_content(r: int, c: int) -> bool:
            """检查指定单元格（0-based索引）是否有内容（有值或属于合并区域）。"""
            if r < 0 or c < 0 or r > max_row or c > max_col:
                return False

            # 1. 检查单元格直接值
            cell = sheet._cells.get((r + 1, c + 1))
            if cell is not None and cell.value is not None:
                return True

            # 2. 检查是否属于某个合并单元格区域
            return merged_lookup.contains_merged_cell(r, c)

        # --- 第一阶段：洪水填充（连通性检测）---
        while queue:
            curr_r, curr_c = queue.popleft()

            # 动态更新表格边界
            min_r = min(min_r, curr_r)
            max_r = max(max_r, curr_r)
            min_c = min(min_c, curr_c)
            max_c = max(max_c, curr_c)

            # 四个方向（上、下、左、右）的邻居检测
            directions = [
                (0, 1),  # 右
                (0, -1),  # 左
                (1, 0),  # 下
                (-1, 0),  # 上
            ]

            for dr, dc in directions:
                # 在容忍距离范围内逐步检查邻居（优先检查最近的）
                for step in range(1, gap_tolerance + 2):
                    nr, nc = curr_r + (dr * step), curr_c + (dc * step)

                    if (nr, nc) in table_cells:
                        break  # 已属于当前表格，不跨越继续查找

                    if has_content(nr, nc):
                        table_cells.add((nr, nc))
                        queue.append((nr, nc))
                        # 在该方向找到连接点，停止扩展间隔
                        break

        # --- 第二阶段：数据提取（语义网格构建）---
        data = []

        # 遍历发现区域的边界框（bbox内部的空格作为空单元格保留，维持矩形布局）
        for ri in range(min_r, max_r + 1):
            for rj in range(min_c, max_c + 1):
                # 跳过被合并单元格遮蔽的单元格（非左上角）
                if merged_lookup.is_hidden_merged_cell(ri, rj):
                    continue

                # 计算合并跨度（默认为 1x1）
                row_span, col_span = merged_lookup.get_anchor_span(ri, rj)

                data.append(
                    self._build_excel_cell(
                        sheet,
                        ri - min_r,  # 相对于表格起始行的偏移
                        rj - min_c,  # 相对于表格起始列的偏移
                        ri,
                        rj,
                        row_span=row_span,
                        col_span=col_span,
                    )
                )

        # 返回给调用方的 visited_cells 严格为包含数据/合并的单元格，
        # 使主循环不会重复扫描已处理的单元格。
        return (
            ExcelTable(
                anchor=(min_c, min_r),
                num_rows=max_r + 1 - min_r,
                num_cols=max_c + 1 - min_c,
                data=data,
            ),
            table_cells,
        )

    def _get_merged_cell_lookup(self, sheet: Worksheet) -> _MergedCellLookup:
        """获取工作表合并单元格缓存，同一轮转换内每个 sheet 只构建一次。"""
        cache_key = id(sheet)
        lookup = self._merged_cell_lookup_cache.get(cache_key)
        if lookup is None:
            lookup = _MergedCellLookup(sheet)
            self._merged_cell_lookup_cache[cache_key] = lookup
        return lookup

    @staticmethod
    def _escape_text_with_line_breaks(text: str) -> str:
        """转义文本并把平台换行统一投影为 HTML 换行。"""
        return html.escape(text).replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")

    @staticmethod
    def _get_cell_hyperlink_target(cell: Any) -> str:
        """读取单元格外链或工作簿内 location。"""
        hyperlink = getattr(cell, "hyperlink", None)
        if not hyperlink:
            return ""

        target = getattr(hyperlink, "target", None)
        if target:
            return str(target)

        location = getattr(hyperlink, "location", None)
        if location:
            return f"#{location}"

        return ""

    @staticmethod
    def _apply_inline_font_tags(text_html: str, inline_font: Any) -> str:
        """按 openpyxl 行内字体顺序包装可见 HTML 标签。"""
        if not text_html or inline_font is None:
            return text_html

        wrapped = text_html
        if getattr(inline_font, "strike", False) or getattr(inline_font, "u", None):
            wrapped = wrapped.replace(" ", "&nbsp;")
        vert_align = getattr(inline_font, "vertAlign", None)
        if vert_align == "superscript":
            wrapped = f"<sup>{wrapped}</sup>"
        elif vert_align == "subscript":
            wrapped = f"<sub>{wrapped}</sub>"

        if getattr(inline_font, "strike", False):
            wrapped = f"<s>{wrapped}</s>"
        if getattr(inline_font, "u", None):
            wrapped = f"<u>{wrapped}</u>"
        if getattr(inline_font, "i", False):
            wrapped = f"<em>{wrapped}</em>"
        if getattr(inline_font, "b", False):
            wrapped = f"<strong>{wrapped}</strong>"

        return wrapped

    def _cell_value_to_html(self, cell: Any) -> tuple[str, bool]:
        """把普通或富文本单元格转换为安全 HTML 与内容类型标记。"""
        if cell.value is None:
            return "", False

        safe_target = sanitize_hyperlink_target(
            self._get_cell_hyperlink_target(cell),
            allowed_schemes=OFFICE_EXTERNAL_HYPERLINK_SCHEMES,
            allow_relative=True,
            allow_fragment=True,
        )
        link_target = html.escape(safe_target, quote=True) if safe_target else ""

        if isinstance(cell.value, CellRichText):
            html_parts = []
            for part in cell.value:
                if hasattr(part, "text"):
                    part_text = self._escape_text_with_line_breaks(str(getattr(part, "text", "")))
                    html_parts.append(
                        self._apply_inline_font_tags(
                            part_text,
                            getattr(part, "font", None),
                        )
                    )
                else:
                    html_parts.append(self._escape_text_with_line_breaks(str(part)))

            rich_text_html = "".join(html_parts)
            if link_target and rich_text_html:
                rich_text_html = f'<a href="{link_target}">{rich_text_html}</a>'
            return rich_text_html, True

        plain_text = str(cell.value)
        if link_target and plain_text:
            escaped_text = self._escape_text_with_line_breaks(plain_text)
            return f'<a href="{link_target}">{escaped_text}</a>', True

        return plain_text, False

    def _extract_cell_style(self, cell: Any) -> dict[str, Any]:
        """从 openpyxl 单元格提取当前 IR 保留的可见样式。"""
        style: dict[str, Any] = {}
        if cell.font:
            if cell.font.b:
                style["font-weight"] = "bold"
            if cell.font.i:
                style["font-style"] = "italic"
            if cell.font.u:
                style["text-decoration"] = "underline"
            if cell.font.strike:
                style["text-decoration"] = "line-through"
            if cell.font.color and hasattr(cell.font.color, "rgb") and cell.font.color.rgb:
                # Color might be ARGB "FF000000"
                color = cell.font.color.rgb
                if isinstance(color, str) and len(color) == 8:
                    style["color"] = "#" + color[2:]
                elif isinstance(color, str):
                    style["color"] = "#" + color

        if cell.alignment:
            if cell.alignment.horizontal:
                style["text-align"] = cell.alignment.horizontal
            if cell.alignment.vertical:
                style["vertical-align"] = cell.alignment.vertical

        if cell.fill and cell.fill.patternType == "solid" and cell.fill.fgColor:
            # handle bg color
            color = cell.fill.fgColor.rgb
            if hasattr(cell.fill.fgColor, "type") and cell.fill.fgColor.type == "rgb" and color:
                if isinstance(color, str) and len(color) == 8:
                    style["background-color"] = "#" + color[2:]
        return style

__all__ = ["SpreadsheetProjector"]
