# Copyright (c) Opendatalab. All rights reserved.

"""把 Excel 97–2003 BIFF 工作簿转换为 MinerU 分页 model-list。"""

from __future__ import annotations

from typing import Any, BinaryIO

from loguru import logger
from openpyxl import Workbook  # type: ignore[reportMissingModuleSource]
from openpyxl.cell.rich_text import CellRichText, TextBlock  # type: ignore[reportMissingModuleSource]
from openpyxl.cell.text import InlineFont  # type: ignore[reportMissingModuleSource]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[reportMissingModuleSource]

from mineru.model.flash.legacy_office import BoundedOleReader, LegacyOfficeEncryptedError
from mineru.model.flash.legacy_office.errors import LegacyOfficeResourceLimitError
from mineru.model.flash.legacy_office.limits import MAX_GRID_SLOTS
from mineru.model.flash.xlsx.xlsx_converter import ExcelTable, XlsxConverter
from mineru.model.office_stream import read_stream_bytes_from_start
from mineru.types import BlockType

from .models import XlsChart, XlsRichText, XlsSheet, XlsWorkbook
from .parser import parse_xls_workbook


def _inline_font(rich_text: XlsRichText, start: int) -> InlineFont | None:
    """返回覆盖指定字符位置的 openpyxl 行内字体。"""

    for run in rich_text.runs:
        if run.start <= start < run.end:
            style = run.style
            return InlineFont(
                b=style.bold,
                i=style.italic,
                u="single" if style.underline else None,
                strike=style.strike,
                vertAlign=(
                    "superscript"
                    if style.superscript
                    else "subscript"
                    if style.subscript
                    else None
                ),
            )
    return None


def _rich_text_boundaries(value: XlsRichText) -> list[int]:
    """收集富文本的起止边界并裁剪到有效字符范围。"""

    boundaries = {0, len(value.text)}
    for run in value.runs:
        boundaries.add(max(0, min(run.start, len(value.text))))
        boundaries.add(max(0, min(run.end, len(value.text))))
    return sorted(boundaries)


def _to_openpyxl_rich_text(value: XlsRichText) -> str | CellRichText:
    """把内部富文本转换为 XlsxConverter 已支持的 CellRichText。"""

    if not value.runs:
        return value.text
    parts: list[str | TextBlock] = []
    boundaries = _rich_text_boundaries(value)
    for start, end in zip(boundaries, boundaries[1:]):
        if start >= end:
            continue
        text = value.text[start:end]
        font = _inline_font(value, start)
        parts.append(TextBlock(font, text) if font is not None else text)
    return CellRichText(parts)


class _XlsPageBuilder(XlsxConverter):
    """用轻量 openpyxl worksheet 适配器复用现有网格与 HTML 投影。"""

    def __init__(self, workbook_model: XlsWorkbook) -> None:
        """初始化解析结果映射和全工作簿网格预算。"""

        super().__init__(include_hidden_sheets=False)
        self.workbook_model = workbook_model
        self._sheet_by_title: dict[str, XlsSheet] = {}
        self._active_xls_sheet: XlsSheet | None = None
        self._grid_slots = 0

    def _build_openpyxl_workbook(self) -> Workbook:
        """把 BIFF 语义模型投影成不落盘的 worksheet 对象。"""

        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)
        for sheet_model in self.workbook_model.sheets:
            worksheet = workbook.create_sheet(sheet_model.name)
            self._sheet_by_title[worksheet.title] = sheet_model
            if not sheet_model.visible:
                worksheet.sheet_state = Worksheet.SHEETSTATE_HIDDEN
            for cell_model in sheet_model.cells.values():
                cell = worksheet.cell(row=cell_model.row + 1, column=cell_model.col + 1)
                cell.value = _to_openpyxl_rich_text(cell_model.value)
                if cell_model.hyperlink:
                    cell.hyperlink = cell_model.hyperlink
            for row_first, col_first, row_last, col_last in sheet_model.merges:
                worksheet.merge_cells(
                    start_row=row_first + 1,
                    start_column=col_first + 1,
                    end_row=row_last + 1,
                    end_column=col_last + 1,
                )
        return workbook

    def build_pages(self) -> list[list[dict[str, Any]]]:
        """逐可见 worksheet 生成 page，并复用 XlsxModel 的标题规则。"""

        self.workbook = self._build_openpyxl_workbook()
        sheet_pages: list[tuple[str, list[dict[str, Any]]]] = []
        for worksheet in self._iter_sheets_to_convert():
            self._active_xls_sheet = self._sheet_by_title.get(worksheet.title)
            self.cur_page = []
            self._convert_sheet(worksheet)
            sheet_pages.append((worksheet.title, self.cur_page))
        if self._should_emit_sheet_titles([page for _, page in sheet_pages]):
            self._prepend_sheet_titles(sheet_pages)
        return [page for _, page in sheet_pages]

    def _collect_sheet_images(self, sheet: Worksheet) -> list[dict]:
        """返回解析器已经绑定到当前 sheet 的图片。"""

        sheet_model = self._sheet_by_title.get(sheet.title)
        if sheet_model is None:
            return []
        return [
            {
                "anchor": (image.row, image.col),
                "base64": image.image_base64,
            }
            for image in sheet_model.images
        ]

    def _find_data_tables(self, sheet: Worksheet) -> list[ExcelTable]:
        """复用 XLSX 区域发现并累计限制实际物化网格。"""

        tables = super()._find_data_tables(sheet)
        self._grid_slots += sum(table.num_rows * table.num_cols for table in tables)
        if self._grid_slots > MAX_GRID_SLOTS:
            raise LegacyOfficeResourceLimitError(
                f"workbook extent exceeds max_grid_slots={MAX_GRID_SLOTS}"
            )
        return tables

    def _chart_block(self, sheet: Worksheet, chart: XlsChart) -> dict[str, Any] | None:
        """把简单 chart 引用转换成数据表，预览图片仅用于无数据 fallback。"""

        if chart.source_rows and chart.source_cols:
            self._grid_slots += len(chart.source_rows) * len(chart.source_cols)
            if self._grid_slots > MAX_GRID_SLOTS:
                raise LegacyOfficeResourceLimitError(
                    f"workbook extent exceeds max_grid_slots={MAX_GRID_SLOTS}"
                )
            table = self._build_synthetic_table_from_sheet_selection(
                sheet,
                list(chart.source_rows),
                list(chart.source_cols),
            )
            return {
                "type": BlockType.CHART,
                "content": self.excel_table_to_html(table),
            }
        if chart.image_base64:
            return {
                "type": BlockType.CHART,
                "content": "",
                "image_base64": chart.image_base64,
            }
        return None

    def _find_charts_in_sheet(
        self,
        sheet: Worksheet,
    ) -> list[tuple[tuple[int, int], int, dict[str, Any]]]:
        """按 cell anchor 输出当前工作表的 legacy chart blocks。"""

        sheet_model = self._sheet_by_title.get(sheet.title)
        if sheet_model is None:
            return []
        artifacts = []
        for order, chart in enumerate(sheet_model.charts):
            block = self._chart_block(sheet, chart)
            if block is None:
                continue
            artifacts.append(((chart.row, chart.col), 10_000 + order, block))
        return artifacts


class XlsConverter:
    """将 Excel 97–2003 OLE/BIFF 二进制流转换为 model-list。"""

    def __init__(self) -> None:
        """初始化空分页输出。"""

        self.pages: list[list[dict[str, Any]]] = []

    def convert(self, file_binary: BinaryIO) -> None:
        """读取 Workbook/Book stream，解析 BIFF 并生成可见工作表页面。"""

        file_bytes = read_stream_bytes_from_start(file_binary)
        with BoundedOleReader(file_bytes) as ole:
            if ole.has_stream("EncryptionInfo") or ole.has_stream("EncryptedPackage"):
                raise LegacyOfficeEncryptedError("password-protected XLS is not supported")
            if ole.has_stream("Workbook"):
                workbook_stream = ole.read_stream("Workbook")
            else:
                workbook_stream = ole.read_stream("Book")
            workbook = parse_xls_workbook(workbook_stream)
        builder = _XlsPageBuilder(workbook)
        self.pages = builder.build_pages()
        logger.debug("XLS parsing produced {} visible worksheet pages", len(self.pages))
