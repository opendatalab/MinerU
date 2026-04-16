# Copyright (c) Opendatalab. All rights reserved.
import collections
import html
import posixpath
import zipfile
import re
import xml.etree.ElementTree as ET
from urllib.parse import urlparse
from typing import BinaryIO, Annotated, cast


from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
from openpyxl.utils.cell import range_to_tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XlsImage
from PIL import Image
from loguru import logger
from pydantic import PositiveInt, Field, BaseModel, NonNegativeInt
from pydantic.dataclasses import dataclass

from mineru.utils.enum_class import BlockType
from mineru.backend.utils.office_image import (
    is_vector_image,
    serialize_vector_image_with_placeholder,
)
from mineru.utils.pdf_reader import image_to_b64str
from mineru.model.docx.tools.math.omml import oMath2Latex

AUTO_GAP_TOLERANCE_CANDIDATES = (0, 1, 2)
AUTO_GAP_TOLERANCE_PREFERENCE = {1: 0, 0: 1, 2: 2}
AUTO_GAP_TOLERANCE_PREFERENCE_MARGIN = 0.15


@dataclass
class DataRegion:
    """表示工作表中非空单元格的边界矩形区域。"""

    min_row: Annotated[
        PositiveInt, Field(description="Smallest row index (1-based index).")
    ]
    max_row: Annotated[
        PositiveInt, Field(description="Largest row index (1-based index).")
    ]
    min_col: Annotated[
        PositiveInt, Field(description="Smallest column index (1-based index).")
    ]
    max_col: Annotated[
        PositiveInt, Field(description="Largest column index (1-based index).")
    ]

    def width(self) -> PositiveInt:
        """返回数据区域的列数。"""
        return self.max_col - self.min_col + 1

    def height(self) -> PositiveInt:
        """返回数据区域的行数。"""
        return self.max_row - self.min_row + 1


class ExcelCell(BaseModel):
    """表示一个 Excel 单元格。

    属性：
        row: 单元格的行号。
        col: 单元格的列号。
        text: 单元格的文本内容。
        row_span: 单元格跨越的行数。
        col_span: 单元格跨越的列数。
    """

    row: int
    col: int
    text: str
    row_span: int
    col_span: int
    styles: dict = Field(default_factory=dict)
    media: list[str] = Field(default_factory=list)
    text_is_html: bool = False
    source_row: int | None = None
    source_col: int | None = None


class ExcelTable(BaseModel):
    """表示工作表上的一个 Excel 表格。

    属性：
        anchor: 表格左上角单元格的列和行索引（从0开始）。
        num_rows: 表格的行数。
        num_cols: 表格的列数。
        data: 表格数据，以 ExcelCell 对象列表的形式表示。
    """

    anchor: tuple[NonNegativeInt, NonNegativeInt]
    num_rows: int
    num_cols: int
    data: list[ExcelCell]


class XlsxConverter:
    def __init__(
        self,
        treat_singleton_as_text: bool = True,
        gap_tolerance: int | None = None,
        include_hidden_sheets: bool = False,
    ):
        self.workbook = None
        self.zf = None
        self.treat_singleton_as_text = treat_singleton_as_text
        self.gap_tolerance = gap_tolerance
        self.include_hidden_sheets = include_hidden_sheets
        self.pages = []
        self.cur_page = []
        self.image_map = {}
        self.cell_image_map = {}
        self.sheet_images = []
        self.table_image_map = {}
        self.math_map = {}
        self.equation_bookends: str = "<eq>{EQ}</eq>"  # 公式标记格式

    def convert(
        self,
        file_stream: BinaryIO,
    ):
        self.pages = []
        self.cur_page = []
        self.sheet_images = []
        self.table_image_map = {}
        self.cell_image_map = {}

        if hasattr(file_stream, "seek"):
            file_stream.seek(0)

        try:
            self.zf = zipfile.ZipFile(file_stream)
        except Exception as e:
            logger.warning(f"Failed to open zip file: {e}")
            self.zf = None

        if hasattr(file_stream, "seek"):
            file_stream.seek(0)

        self.workbook = load_workbook(
            filename=file_stream,
            data_only=True,
            rich_text=True,
        )
        if self.workbook is not None:
            # 遍历需要参与转换的工作表，避免为隐藏表或尾部空页生成无效页面。
            for idx, sheet in enumerate(self._iter_sheets_to_convert(), start=1):
                logger.debug(f"正在处理第 {idx} 个工作表：{sheet.title}")
                self.cur_page = []
                self._convert_sheet(sheet)
                self.pages.append(self.cur_page)
        else:
            logger.error("工作簿未初始化。")

        if self.zf:
            self.zf.close()
            self.zf = None

    def _iter_sheets_to_convert(self):
        if self.workbook is None:
            return

        for sheet in self.workbook.worksheets:
            if (
                not self.include_hidden_sheets
                and sheet.sheet_state != Worksheet.SHEETSTATE_VISIBLE
            ):
                logger.debug(f"跳过隐藏工作表：{sheet.title}")
                continue
            yield sheet

    def _convert_sheet(self, sheet):
        if isinstance(sheet, Worksheet):
            # Pre-calc maps
            self.math_map = self._map_math_formulas_to_cells(sheet)
            self.sheet_images = self._collect_sheet_images(sheet)
            self.table_image_map = collections.defaultdict(list)
            for image_info in self.sheet_images:
                anchor = image_info["anchor"]
                if anchor[0] is None or anchor[1] is None:
                    continue
                self.table_image_map[anchor].append(
                    f'<img src="{image_info["base64"]}" />'
                )

            used_cells, visual_artifacts = self._find_tables_in_sheet(sheet)
            visual_artifacts.extend(self._find_charts_in_sheet(sheet))
            for _, _, block in sorted(
                visual_artifacts,
                key=lambda item: (item[0][0], item[0][1], item[1]),
            ):
                self.cur_page.append(block)
            self._find_images_in_sheet(used_cells)  # 提取图片

    @staticmethod
    def _serialize_sheet_image(image: XlsImage) -> str:
        pil_image = Image.open(image.ref)  # type: ignore[arg-type]
        if is_vector_image(pil_image):
            return serialize_vector_image_with_placeholder(pil_image)

        if pil_image.mode != "RGB":
            return image_to_b64str(pil_image, image_format="PNG")

        return image_to_b64str(pil_image, image_format="JPEG")

    def _collect_sheet_images(self, sheet: Worksheet) -> list[dict]:
        images = []
        if self.workbook is None:
            return images

        for item in getattr(sheet, "_images", []):  # type: ignore[attr-defined]
            try:
                image: XlsImage = cast(XlsImage, item)
                images.append(
                    {
                        "anchor": self._get_anchor_pos(item.anchor),
                        "base64": self._serialize_sheet_image(image),
                    }
                )
            except Exception as e:
                logger.error(f"无法从 Excel 工作表中提取图片，错误信息：{e}")

        return images

    def _map_math_formulas_to_cells(self, sheet: Worksheet) -> dict:
        """Parse drawings to find math formulas and map them to cells."""
        math_map = collections.defaultdict(list)
        if not self.zf:
            return math_map

        # Find drawing relation
        drawing_rel = None
        if hasattr(sheet, "_rels"):
            for rel in sheet._rels:
                if rel.Type.endswith("/relationships/drawing"):
                    drawing_rel = rel
                    break

        if not drawing_rel:
            return math_map

        # Resolve path
        # Assuming relative path from worksheets/sheetX.xml to drawings/drawingY.xml
        # Usually target is like "../drawings/drawing1.xml"
        target = drawing_rel.Target
        if target.startswith("../"):
            path = target.replace("../", "xl/")  # simplistic resolution
        elif target.startswith("/"):
            path = target[1:]
        else:
            path = f"xl/worksheets/{target}"  # unlikely but default relative

        # Check if file exists in zip
        if path not in self.zf.namelist():
            # Try generic match if simplistic resolution failed
            # drawing1.xml -> xl/drawings/drawing1.xml
            basename = target.split("/")[-1]
            path = f"xl/drawings/{basename}"
            if path not in self.zf.namelist():
                return math_map

        try:
            with self.zf.open(path) as f:
                tree = ET.parse(f)
                root = tree.getroot()

            # Namespaces
            ns = {
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                "m": "http://schemas.openxmlformats.org/officeDocument/2006/math",
            }

            # Iterate TwoCellAnchor and OneCellAnchor
            for anchor_tag in ["twoCellAnchor", "oneCellAnchor"]:
                for anchor in root.findall(f".//xdr:{anchor_tag}", ns):
                    # Get position
                    from_node = anchor.find("xdr:from", ns)
                    if from_node is None:
                        continue
                    col_node = from_node.find("xdr:col", ns)
                    row_node = from_node.find("xdr:row", ns)
                    if col_node is None or row_node is None:
                        continue

                    r = int(row_node.text)
                    c = int(col_node.text)

                    # Look for math content
                    # Usually in graphicalFrame -> graphic -> graphicData -> oMathPara
                    # But simpler to search descendant m:oMath
                    maths = anchor.findall(".//m:oMath", ns)
                    for math in maths:
                        # # Simple text extraction
                        # text = "".join(math.itertext())
                        # if text.strip():
                        #     # Wrap in latex block indicator if needed, or just plain text
                        #     # User asked for formula, assuming latex-like visual or text is acceptable
                        #     # Adding simple latex-like wrapper
                        #     math_map[(r, c)].append(f"${text}$")
                        latex = str(oMath2Latex(math)).strip()
                        if latex:
                            math_map[(r, c)].append(latex)

        except Exception as e:
            logger.warning(f"Error parsing math formulas: {e}")

        return math_map

    def _get_anchor_pos(self, anchor):
        """Helper to get (row, col) from anchor."""
        if hasattr(anchor, "_from"):
            return anchor._from.row, anchor._from.col
        return None, None

    def _get_block_sort_anchor(
        self, row: int | None, col: int | None
    ) -> tuple[int, int]:
        if row is None or col is None:
            return (10**9, 10**9)
        return row, col

    def _build_block_from_excel_table(self, excel_table: ExcelTable) -> dict:
        if (
            self.treat_singleton_as_text
            and len(excel_table.data) == 1
            and self._can_render_singleton_as_text(excel_table)
        ):
            return {
                "type": BlockType.TEXT,
                "content": excel_table.data[0].text,
            }

        return {
            "type": BlockType.TABLE,
            "content": self.excel_table_to_html(excel_table),
        }

    def _find_tables_in_sheet(
        self, sheet: Worksheet
    ) -> tuple[set[tuple[int, int]], list[tuple[tuple[int, int], int, dict]]]:
        used_cells = set()
        visual_artifacts = []
        if self.workbook is not None:
            content_layer = self._get_sheet_content_layer(sheet)  # 检测工作表的可见性
            tables = self._find_data_tables(sheet)  # 检测工作表中的所有数据表格

            for order, excel_table in enumerate(tables):
                # Record used cells
                anchor_c, anchor_r = excel_table.anchor
                for cell in excel_table.data:
                    source_row, source_col = self._resolve_excel_cell_source_position(
                        excel_table.anchor,
                        cell,
                    )
                    used_cells.add((source_row, source_col))

                visual_artifacts.append(
                    (
                        self._get_block_sort_anchor(anchor_r, anchor_c),
                        order,
                        self._build_block_from_excel_table(excel_table),
                    )
                )

        return used_cells, visual_artifacts

    def _extract_chart_range_formula(self, value_source) -> str | None:
        if value_source is None:
            return None

        for attr_name in ("numRef", "strRef", "multiLvlStrRef"):
            ref = getattr(value_source, attr_name, None)
            formula = getattr(ref, "f", None)
            if formula:
                return formula

        return None

    def _iter_chart_reference_formulas(self, chart):
        for series in getattr(chart, "ser", []):
            for attr_name in ("cat", "val", "xVal", "yVal", "bubbleSize"):
                formula = self._extract_chart_range_formula(
                    getattr(series, attr_name, None)
                )
                if formula:
                    yield formula

            tx = getattr(series, "tx", None)
            tx_formula = getattr(getattr(tx, "strRef", None), "f", None)
            if tx_formula:
                yield tx_formula

    def _parse_chart_reference_formula(
        self, formula: str, sheet_title: str
    ) -> tuple[list[int], list[int]] | None:
        try:
            formula_sheet_name, (
                min_col,
                min_row,
                max_col,
                max_row,
            ) = range_to_tuple(formula)
        except ValueError:
            logger.debug("Skip unsupported chart reference formula: {}", formula)
            return None

        if formula_sheet_name != sheet_title:
            logger.debug(
                "Skip chart reference formula from different sheet: {} != {}",
                formula_sheet_name,
                sheet_title,
            )
            return None

        if not all(
            isinstance(bound, int)
            for bound in (min_col, min_row, max_col, max_row)
        ):
            logger.debug(
                "Skip chart reference formula with open-ended bounds: {}",
                formula,
            )
            return None

        rows = list(range(min_row - 1, max_row))
        cols = list(range(min_col - 1, max_col))
        return rows, cols

    def _collect_chart_source_axes(
        self, sheet: Worksheet, chart
    ) -> tuple[list[int], list[int]] | None:
        referenced_rows = set()
        referenced_cols = set()
        formulas_found = False

        for formula in self._iter_chart_reference_formulas(chart):
            formulas_found = True
            parsed_axes = self._parse_chart_reference_formula(formula, sheet.title)
            if parsed_axes is None:
                return None

            rows, cols = parsed_axes
            referenced_rows.update(rows)
            referenced_cols.update(cols)

        if not formulas_found or not referenced_rows or not referenced_cols:
            return None

        return sorted(referenced_rows), sorted(referenced_cols)

    def _build_excel_cell(
        self,
        sheet: Worksheet,
        display_row: int,
        display_col: int,
        source_row: int,
        source_col: int,
        row_span: int = 1,
        col_span: int = 1,
    ) -> ExcelCell:
        cell = sheet.cell(row=source_row + 1, column=source_col + 1)
        raw_cell_text = str(cell.value) if cell.value is not None else ""
        cell_text = ""
        text_is_html = False
        media_content = []
        if "DISPIMG" in raw_cell_text:
            cell_image = self._get_cell_image(raw_cell_text)
            if cell_image:
                media_content.append(cell_image)
        else:
            cell_text, text_is_html = self._cell_value_to_html(cell)
        media_content.extend(self.table_image_map.get((source_row, source_col), []))

        return ExcelCell(
            row=display_row,
            col=display_col,
            text=cell_text,
            row_span=row_span,
            col_span=col_span,
            styles=self._extract_cell_style(cell),
            media=media_content,
            text_is_html=text_is_html,
            source_row=source_row,
            source_col=source_col,
        )

    def _build_synthetic_table_from_sheet_selection(
        self, sheet: Worksheet, rows: list[int], cols: list[int]
    ) -> ExcelTable:
        selected_coords = {(row, col) for row in rows for col in cols}
        hidden_merge_cells = set()
        merge_spans = {}

        for mr in sheet.merged_cells.ranges:
            top_left = (mr.min_row - 1, mr.min_col - 1)
            if top_left not in selected_coords:
                continue

            selected_rows = [
                row for row in rows if mr.min_row - 1 <= row <= mr.max_row - 1
            ]
            selected_cols = [
                col for col in cols if mr.min_col - 1 <= col <= mr.max_col - 1
            ]
            if not selected_rows or not selected_cols:
                continue

            merge_spans[top_left] = (len(selected_rows), len(selected_cols))
            for row in selected_rows:
                for col in selected_cols:
                    if (row, col) != top_left:
                        hidden_merge_cells.add((row, col))

        data = []
        for display_row, source_row in enumerate(rows):
            for display_col, source_col in enumerate(cols):
                if (source_row, source_col) in hidden_merge_cells:
                    continue

                row_span, col_span = merge_spans.get((source_row, source_col), (1, 1))
                data.append(
                    self._build_excel_cell(
                        sheet,
                        display_row,
                        display_col,
                        source_row,
                        source_col,
                        row_span=row_span,
                        col_span=col_span,
                    )
                )

        return ExcelTable(
            anchor=(cols[0], rows[0]),
            num_rows=len(rows),
            num_cols=len(cols),
            data=data,
        )

    def _find_charts_in_sheet(
        self, sheet: Worksheet
    ) -> list[tuple[tuple[int, int], int, dict]]:
        chart_artifacts = []
        for order, chart in enumerate(getattr(sheet, "_charts", [])):
            axes = self._collect_chart_source_axes(sheet, chart)
            if axes is None:
                logger.debug(
                    "Skip chart on sheet '{}' because chart source ranges are unsupported",
                    sheet.title,
                )
                continue

            rows, cols = axes
            chart_table = self._build_synthetic_table_from_sheet_selection(
                sheet,
                rows,
                cols,
            )
            anchor_row, anchor_col = self._get_anchor_pos(getattr(chart, "anchor", None))
            chart_artifacts.append(
                (
                    self._get_block_sort_anchor(anchor_row, anchor_col),
                    10_000 + order,
                    {
                        "type": BlockType.CHART,
                        "content": self.excel_table_to_html(chart_table),
                    },
                )
            )

        return chart_artifacts

    def _get_cell_math_formulas(
        self,
        table_anchor: tuple[int, int],
        row: int | None = None,
        col: int | None = None,
        excel_cell: ExcelCell | None = None,
    ) -> list[str]:
        abs_row, abs_col = self._resolve_excel_cell_source_position(
            table_anchor,
            excel_cell,
            row=row,
            col=col,
        )
        return list(self.math_map.get((abs_row, abs_col), []))

    def _resolve_excel_cell_source_position(
        self,
        table_anchor: tuple[int, int],
        excel_cell: ExcelCell | None,
        row: int | None = None,
        col: int | None = None,
    ) -> tuple[int, int]:
        if excel_cell is not None:
            if excel_cell.source_row is not None and excel_cell.source_col is not None:
                return excel_cell.source_row, excel_cell.source_col
            row = excel_cell.row
            col = excel_cell.col

        if row is None or col is None:
            raise ValueError("row and col must be provided when excel_cell is None")

        return table_anchor[1] + row, table_anchor[0] + col

    def _can_render_singleton_as_text(self, excel_table: ExcelTable) -> bool:
        cell = excel_table.data[0]
        return (
            cell.row_span == 1
            and cell.col_span == 1
            and not cell.media
            and not cell.text_is_html
            and not self._get_cell_math_formulas(
                excel_table.anchor,
                excel_cell=cell,
            )
        )

    def _cell_has_semantic_content(
        self, excel_table: ExcelTable, cell: ExcelCell
    ) -> bool:
        return bool(
            cell.text.strip()
            or any(media.strip() for media in cell.media)
            or self._get_cell_math_formulas(excel_table.anchor, excel_cell=cell)
        )

    def _get_table_semantic_positions(
        self, excel_table: ExcelTable
    ) -> set[tuple[int, int]]:
        semantic_positions = set()
        for cell in excel_table.data:
            if not self._cell_has_semantic_content(excel_table, cell):
                continue
            semantic_positions.add(
                self._resolve_excel_cell_source_position(
                    excel_table.anchor,
                    excel_cell=cell,
                )
            )
        return semantic_positions

    def _filter_semantic_subset_tables(
        self, tables: list[ExcelTable]
    ) -> list[ExcelTable]:
        semantic_positions = [
            self._get_table_semantic_positions(table) for table in tables
        ]
        filtered_tables = []

        for table_idx, table in enumerate(tables):
            if any(
                semantic_positions[table_idx] < semantic_positions[other_idx]
                for other_idx in range(len(tables))
                if other_idx != table_idx
            ):
                continue
            filtered_tables.append(table)

        return filtered_tables

    def _build_table_content_mask(self, excel_table: ExcelTable) -> list[list[bool]]:
        mask = [
            [False for _ in range(excel_table.num_cols)]
            for _ in range(excel_table.num_rows)
        ]
        for cell in excel_table.data:
            if not self._cell_has_semantic_content(excel_table, cell):
                continue
            for row_idx in range(cell.row, min(cell.row + cell.row_span, excel_table.num_rows)):
                for col_idx in range(
                    cell.col, min(cell.col + cell.col_span, excel_table.num_cols)
                ):
                    mask[row_idx][col_idx] = True
        return mask

    @staticmethod
    def _count_max_consecutive_true(flags: list[bool]) -> int:
        max_count = 0
        current = 0
        for flag in flags:
            if flag:
                current += 1
                max_count = max(max_count, current)
            else:
                current = 0
        return max_count

    @staticmethod
    def _is_real_singleton_table(excel_table: ExcelTable) -> bool:
        if (
            excel_table.num_rows != 1
            or excel_table.num_cols != 1
            or len(excel_table.data) != 1
        ):
            return False
        cell = excel_table.data[0]
        return cell.row_span == 1 and cell.col_span == 1

    def _summarize_table_for_gap_selection(
        self, excel_table: ExcelTable
    ) -> dict[str, float | int | bool]:
        table_area = excel_table.num_rows * excel_table.num_cols
        content_mask = self._build_table_content_mask(excel_table)
        content_area = sum(sum(1 for flag in row if flag) for row in content_mask)
        blank_ratio = 1.0 - (content_area / max(table_area, 1))

        interior_blank_rows = [
            not any(content_mask[row_idx])
            for row_idx in range(1, max(excel_table.num_rows - 1, 1))
        ]
        interior_blank_cols = [
            not any(content_mask[row_idx][col_idx] for row_idx in range(excel_table.num_rows))
            for col_idx in range(1, max(excel_table.num_cols - 1, 1))
        ]
        if excel_table.num_rows <= 2:
            interior_blank_rows = []
        if excel_table.num_cols <= 2:
            interior_blank_cols = []

        interior_blank_row_count = sum(interior_blank_rows)
        interior_blank_col_count = sum(interior_blank_cols)
        max_consecutive_interior_blank_lines = max(
            self._count_max_consecutive_true(interior_blank_rows),
            self._count_max_consecutive_true(interior_blank_cols),
        )

        return {
            "table_area": table_area,
            "content_area": content_area,
            "blank_ratio": blank_ratio,
            "interior_blank_row_count": interior_blank_row_count,
            "interior_blank_col_count": interior_blank_col_count,
            "max_consecutive_interior_blank_lines": max_consecutive_interior_blank_lines,
            "real_singleton": self._is_real_singleton_table(excel_table),
        }

    def _summarize_candidate_tables(
        self, tables: list[ExcelTable]
    ) -> dict[str, float | int]:
        table_count = len(tables)
        real_singleton_count = 0
        severe_separator_count = 0
        sparse_large_table_count = 0
        total_area = 0
        weighted_blank_numerator = 0.0
        total_interior_blank_lines = 0
        total_possible_interior_lines = 0
        row_cover_count = collections.Counter()

        for table in tables:
            table_summary = self._summarize_table_for_gap_selection(table)
            table_area = int(table_summary["table_area"])
            blank_ratio = float(table_summary["blank_ratio"])
            interior_blank_row_count = int(table_summary["interior_blank_row_count"])
            interior_blank_col_count = int(table_summary["interior_blank_col_count"])
            max_consecutive_interior_blank_lines = int(
                table_summary["max_consecutive_interior_blank_lines"]
            )

            total_area += table_area
            weighted_blank_numerator += table_area * blank_ratio
            total_interior_blank_lines += (
                interior_blank_row_count + interior_blank_col_count
            )
            total_possible_interior_lines += max(table.num_rows - 2, 0) + max(
                table.num_cols - 2, 0
            )
            for row_idx in range(table.anchor[1], table.anchor[1] + table.num_rows):
                row_cover_count[row_idx] += 1

            if bool(table_summary["real_singleton"]):
                real_singleton_count += 1
            if table_area >= 6 and blank_ratio > 0.35:
                sparse_large_table_count += 1
            if max_consecutive_interior_blank_lines >= 2:
                severe_separator_count += 1

        occupied_row_count = max(len(row_cover_count), 1)
        row_overlap_excess_ratio = sum(
            max(0, count - 1) for count in row_cover_count.values()
        ) / occupied_row_count

        return {
            "real_singleton_ratio": real_singleton_count / max(table_count, 1),
            "weighted_blank_ratio": weighted_blank_numerator / max(total_area, 1),
            "interior_blank_line_ratio": total_interior_blank_lines
            / max(total_possible_interior_lines, 1),
            "sparse_large_table_ratio": sparse_large_table_count / max(table_count, 1),
            "severe_separator_count": severe_separator_count,
            "row_overlap_excess_ratio": row_overlap_excess_ratio,
        }

    def _select_best_gap_candidate(
        self, sheet: Worksheet
    ) -> tuple[int, float, list[ExcelTable]]:
        candidates = []
        for gap_tolerance in AUTO_GAP_TOLERANCE_CANDIDATES:
            raw_tables = self._find_data_tables_with_gap_raw(sheet, gap_tolerance)
            summary = self._summarize_candidate_tables(raw_tables)
            penalty = (
                6.0 * int(summary["severe_separator_count"])
                + 2.5 * float(summary["interior_blank_line_ratio"])
                + 1.5 * float(summary["sparse_large_table_ratio"])
                + 1.0 * float(summary["real_singleton_ratio"])
                + 0.5 * float(summary["weighted_blank_ratio"])
                + 1.0 * float(summary["row_overlap_excess_ratio"])
            )
            candidates.append(
                {
                    "gap_tolerance": gap_tolerance,
                    "penalty": penalty,
                    "tables": self._filter_semantic_subset_tables(raw_tables),
                    **summary,
                }
            )

        min_penalty = min(float(candidate["penalty"]) for candidate in candidates)
        near_best_candidates = [
            candidate
            for candidate in candidates
            if float(candidate["penalty"])
            <= (min_penalty + AUTO_GAP_TOLERANCE_PREFERENCE_MARGIN)
        ]

        best_candidate = min(
            near_best_candidates,
            key=lambda candidate: (
                int(candidate["severe_separator_count"]),
                AUTO_GAP_TOLERANCE_PREFERENCE[int(candidate["gap_tolerance"])],
                float(candidate["interior_blank_line_ratio"]),
                float(candidate["penalty"]),
            ),
        )
        return (
            int(best_candidate["gap_tolerance"]),
            float(best_candidate["penalty"]),
            best_candidate["tables"],
        )

    def _select_best_tables(self, sheet: Worksheet) -> list[ExcelTable]:
        gap_tolerance, penalty, tables = self._select_best_gap_candidate(sheet)
        logger.debug(
            "Selected gap_tolerance={} for sheet '{}' with penalty={:.4f}",
            gap_tolerance,
            sheet.title,
            penalty,
        )
        return tables

    def excel_table_to_html(self, excel_table) -> str:
        """
        将 ExcelTable 转换为 HTML 表格字符串，保留合并单元格结构。
        """
        # 1. 创建坐标到单元格的映射，方便快速查找
        cell_map = {(c.row, c.col): c for c in excel_table.data}
        table_anchor = excel_table.anchor

        # 2. 用于记录已被合并单元格占据的位置，避免重复生成 td
        covered_cells = set()

        # 开始构建 HTML
        lines = ["<table>"]  # 可以根据需要添加样式类或属性

        for r in range(excel_table.num_rows):
            lines.append("  <tr>")
            for c in range(excel_table.num_cols):
                # 如果当前位置已被之前的合并单元格占据，则跳过
                if (r, c) in covered_cells:
                    continue

                # 获取当前位置的单元格
                cell = cell_map.get((r, c))

                if cell:
                    # 确定标签类型：第一行通常作为表头
                    tag = "th" if cell.row == 0 else "td"

                    # 构建属性列表 (rowspan, colspan)
                    attrs = []
                    if cell.row_span > 1:
                        attrs.append(f'rowspan="{cell.row_span}"')
                    if cell.col_span > 1:
                        attrs.append(f'colspan="{cell.col_span}"')

                    # 标记该单元格覆盖的所有位置为已占用
                    for ir in range(cell.row_span):
                        for ic in range(cell.col_span):
                            covered_cells.add((r + ir, c + ic))

                    # 拼接属性字符串
                    attr_str = " " + " ".join(attrs) if attrs else ""

                    # 生成 HTML 单元格，富文本片段避免二次转义
                    text_content = ""
                    if cell.text:
                        text_content = cell.text if cell.text_is_html else html.escape(cell.text)

                    # 添加媒体内容 (Images)
                    if cell.media:
                        media_content = "<br>".join(cell.media)
                        if text_content:
                            text_content += "<br>" + media_content
                        else:
                            text_content = media_content
                    # 添加公式
                    for formula in self._get_cell_math_formulas(
                        table_anchor,
                        excel_cell=cell,
                    ):
                        text_content += self.equation_bookends.format(EQ=formula)

                    inner_html = self._render_cell_inner_html(
                        text_content,
                        cell.text_is_html,
                    )
                    lines.append(f"    <{tag}{attr_str}>{inner_html}</{tag}>")
                else:
                    # 如果既没被覆盖，又没有数据对象（理论上 _find_table_bounds 逻辑应避免此情况），生成空单元格
                    lines.append("    <td></td>")

            lines.append("  </tr>")

        lines.append("</table>")
        return "\n".join(lines)

    def _find_images_in_sheet(self, used_cells: set[tuple[int, int]] = None):
        if self.workbook is not None:
            for image_info in self.sheet_images:
                r, c = image_info["anchor"]
                if (
                    used_cells
                    and r is not None
                    and c is not None
                    and (r, c) in used_cells
                ):
                    continue

                self.cur_page.append(
                    {
                        "type": BlockType.IMAGE,
                        "content": image_info["base64"],
                    }
                )

        return

    def _find_data_tables(self, sheet: Worksheet) -> list[ExcelTable]:
        """在 Excel 工作表中查找所有紧凑的矩形数据表格。

        参数：
            sheet: 待解析的 Excel 工作表。

        返回：
            表示所有数据表格的 ExcelTable 对象列表。
        """
        if self.gap_tolerance is None:
            return self._select_best_tables(sheet)
        return self._find_data_tables_with_gap(sheet, self.gap_tolerance)

    def _find_data_tables_with_gap(
        self, sheet: Worksheet, gap_tolerance: int
    ) -> list[ExcelTable]:
        return self._filter_semantic_subset_tables(
            self._find_data_tables_with_gap_raw(sheet, gap_tolerance)
        )

    def _find_data_tables_with_gap_raw(
        self, sheet: Worksheet, gap_tolerance: int
    ) -> list[ExcelTable]:
        """在固定 gap_tolerance 下查找工作表中的所有数据表格。"""
        bounds: DataRegion = self._find_true_data_bounds(sheet)  # 获取真实数据边界
        tables: list[ExcelTable] = []  # 存储已发现的表格
        visited: set[tuple[int, int]] = set()  # 记录已访问的单元格

        # 仅在真实数据边界范围内进行扫描
        for ri, row in enumerate(
            sheet.iter_rows(
                min_row=bounds.min_row,
                max_row=bounds.max_row,
                min_col=bounds.min_col,
                max_col=bounds.max_col,
                values_only=False,
            ),
            start=bounds.min_row - 1,  # 转换为0-based索引
        ):
            for rj, cell in enumerate(row, start=bounds.min_col - 1):
                # 跳过空单元格或已访问的单元格
                if cell.value is None or (ri, rj) in visited:
                    continue

                # 从当前单元格出发，通过洪水填充算法确定所属表格的边界
                table_bounds, visited_cells = self._find_table_bounds(
                    sheet,
                    ri,
                    rj,
                    bounds.max_row,
                    bounds.max_col,
                    gap_tolerance,
                )
                visited.update(visited_cells)  # 将已访问单元格加入全局记录
                tables.append(table_bounds)

        return tables

    def _find_true_data_bounds(self, sheet: Worksheet) -> DataRegion:
        """查找工作表中真实的数据边界（最小/最大行列）。

        该函数扫描所有单元格，找到包含所有非空单元格或合并单元格区域的
        最小矩形范围，返回边界的行列索引。

        参数：
            sheet: 待分析的工作表。

        返回：
            覆盖所有数据和合并单元格的最小矩形区域 DataRegion。
            若工作表为空，则默认返回 (1, 1, 1, 1)。
        """
        min_row, min_col = None, None
        max_row, max_col = 0, 0

        # 遍历所有有值的单元格，动态更新边界
        for cell in sheet._cells.values():
            if cell.value is not None:
                r, c = cell.row, cell.column
                min_row = r if min_row is None else min(min_row, r)
                min_col = c if min_col is None else min(min_col, c)
                max_row = max(max_row, r)
                max_col = max(max_col, c)

        # 将合并单元格的范围也纳入边界计算
        for merged in sheet.merged_cells.ranges:
            min_row = (
                merged.min_row if min_row is None else min(min_row, merged.min_row)
            )
            min_col = (
                merged.min_col if min_col is None else min(min_col, merged.min_col)
            )
            max_row = max(max_row, merged.max_row)
            max_col = max(max_col, merged.max_col)

        # 若工作表中没有任何数据，默认返回 (1, 1, 1, 1)
        if min_row is None or min_col is None:
            min_row = min_col = max_row = max_col = 1

        return DataRegion(min_row, max_row, min_col, max_col)

    def _find_table_bounds(
        self,
        sheet: Worksheet,
        start_row: int,
        start_col: int,
        max_row: int,
        max_col: int,
        gap_tolerance: int,
    ) -> tuple[ExcelTable, set[tuple[int, int]]]:
        """使用洪水填充（BFS）策略确定表格边界。

        该方法通过广度优先搜索（BFS）算法识别 Excel 工作表中连续的非空单元格区域，
        能够准确检测非矩形表格（如 L 形、错位列等），并支持通过间隔容忍度
        连接相邻但不直接相连的单元格。

        算法分两个阶段执行：
        1. 洪水填充阶段：使用 BFS 从给定位置出发，找出所有相连的单元格。
        2. 数据提取阶段：构建矩形边界框并提取单元格数据，正确处理合并单元格。

        参数：
            sheet: 待分析的 Excel 工作表。
            start_row: 洪水填充起始行索引（从0开始）。
            start_col: 洪水填充起始列索引（从0开始）。
            max_row: 工作表中可考虑的最大行索引（从0开始）。
            max_col: 工作表中可考虑的最大列索引（从0开始）。
            gap_tolerance: 允许跨越空白单元格查找邻居的最大间隔。

        返回：
            一个元组，包含：
                - ExcelTable：表示检测到的表格对象，含锚点位置、尺寸和单元格数据。
                - set[tuple[int, int]]：洪水填充期间访问的所有 (行, 列) 元组集合，
                  用于防止重复扫描。

        说明：
            该方法遵循 GAP_TOLERANCE 选项，允许在容忍距离内将被空单元格隔开的
            单元格视为同一表格的一部分。
        """

        # BFS 队列，存储待处理的 (行, 列) 坐标
        queue = collections.deque([(start_row, start_col)])

        # 记录当前表格内已访问的单元格（避免重复加入队列）
        # 调用方维护全局 visited 集合，防止重复启动新表格
        table_cells: set[tuple[int, int]] = set()
        table_cells.add((start_row, start_col))

        # 动态记录当前表格的行列边界
        min_r, max_r = start_row, start_row
        min_c, max_c = start_col, start_col

        def has_content(r, c):
            """检查指定单元格（0-based索引）是否有内容（有值或属于合并区域）。"""
            if r < 0 or c < 0 or r > max_row or c > max_col:
                return False

            # 1. 检查单元格直接值
            cell = sheet.cell(row=r + 1, column=c + 1)
            if cell.value is not None:
                return True

            # 2. 检查是否属于某个合并单元格区域
            for mr in sheet.merged_cells.ranges:
                if cell.coordinate in mr:
                    return True
            return False

        # --- 第一阶段：洪水填充（连通性检测）---
        while queue:
            curr_r, curr_c = queue.popleft()

            # 动态更新表格边界
            min_r = min(min_r, curr_r)
            max_r = max(max_r, curr_r)
            min_c = min(min_c, curr_c)
            max_c = max(max_c, curr_c)

            # 四个方向（上、下、左、右）的邻居检测
            directions = [
                (0, 1),  # 右
                (0, -1),  # 左
                (1, 0),  # 下
                (-1, 0),  # 上
            ]

            for dr, dc in directions:
                # 在容忍距离范围内逐步检查邻居（优先检查最近的）
                for step in range(1, gap_tolerance + 2):
                    nr, nc = curr_r + (dr * step), curr_c + (dc * step)

                    if (nr, nc) in table_cells:
                        break  # 已属于当前表格，不跨越继续查找

                    if has_content(nr, nc):
                        table_cells.add((nr, nc))
                        queue.append((nr, nc))
                        # 在该方向找到连接点，停止扩展间隔
                        break

        # --- 第二阶段：数据提取（语义网格构建）---
        data = []

        # 识别被合并单元格"遮蔽"的单元格（即非合并区域左上角的单元格）
        hidden_merge_cells = set()
        for mr in sheet.merged_cells.ranges:
            mr_min_r, mr_min_c = mr.min_row - 1, mr.min_col - 1
            mr_max_r, mr_max_c = mr.max_row - 1, mr.max_col - 1
            for r in range(mr_min_r, mr_max_r + 1):
                for c in range(mr_min_c, mr_max_c + 1):
                    if r == mr_min_r and c == mr_min_c:
                        continue  # 左上角单元格保留，其余标记为隐藏
                    hidden_merge_cells.add((r, c))

        # 遍历发现区域的边界框（bbox内部的空格作为空单元格保留，维持矩形布局）
        for ri in range(min_r, max_r + 1):
            for rj in range(min_c, max_c + 1):
                # 跳过被合并单元格遮蔽的单元格（非左上角）
                if (ri, rj) in hidden_merge_cells:
                    continue

                # 计算合并跨度（默认为 1x1）
                row_span = 1
                col_span = 1
                for mr in sheet.merged_cells.ranges:
                    if (ri + 1) == mr.min_row and (rj + 1) == mr.min_col:
                        row_span = (mr.max_row - mr.min_row) + 1
                        col_span = (mr.max_col - mr.min_col) + 1
                        break

                data.append(
                    self._build_excel_cell(
                        sheet,
                        ri - min_r,  # 相对于表格起始行的偏移
                        rj - min_c,  # 相对于表格起始列的偏移
                        ri,
                        rj,
                        row_span=row_span,
                        col_span=col_span,
                    )
                )

        # 返回给调用方的 visited_cells 严格为包含数据/合并的单元格，
        # 使主循环不会重复扫描已处理的单元格。
        return (
            ExcelTable(
                anchor=(min_c, min_r),
                num_rows=max_r + 1 - min_r,
                num_cols=max_c + 1 - min_c,
                data=data,
            ),
            table_cells,
        )

    def _get_cell_image(self, text) -> str:
        match = re.search(r'"([^"]+)"', text)
        if match:
            image_id = match.group(1)

        else:
            logger.error(f"无法从单元格文本中提取图片 ID，文本内容：{text}")
            return ""

        cell_image_map = self._load_cell_image_mappings()

        zip_target_path = posixpath.normpath(posixpath.join("xl", cell_image_map.get(image_id, "")))
        if self.zf is None or zip_target_path not in self.zf.namelist():
            logger.warning(
                f"图片目标文件不存在，image_id={image_id}, target={zip_target_path}"
            )
            return ""

        try:
            with self.zf.open(zip_target_path) as image_file:
                pil_image = Image.open(image_file)
                if is_vector_image(pil_image):
                    img_base64 = serialize_vector_image_with_placeholder(pil_image)
                    return rf'<img src="{img_base64}" />'

                pil_image.load()

                if pil_image.mode != "RGB":
                    img_base64 = image_to_b64str(pil_image, image_format="PNG")
                else:
                    img_base64 = image_to_b64str(pil_image, image_format="JPEG")
                return rf'<img src="{img_base64}" />'
        except Exception as e:
            logger.warning(
                f"读取单元格图片失败，image_id={image_id}, target={zip_target_path}, error={e}"
            )
            return ""

    def _load_cell_image_mappings(self):
        if self.cell_image_map:
            return self.cell_image_map

        if self.zf is None:
            return {}
        cell_image_embed_to_name = {}
        cellimages_path = "xl/cellimages.xml"
        rels_path = "xl/_rels/cellimages.xml.rels"
        if (
            cellimages_path not in self.zf.namelist()
            or rels_path not in self.zf.namelist()
        ):
            return {}

        try:
            with self.zf.open(cellimages_path) as f:
                root = ET.parse(f).getroot()

            ns = {
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
                "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
            }

            for cell_image in root.findall(".//etc:cellImage", ns):
                c_nv_pr = cell_image.find(".//xdr:cNvPr", ns)
                blip = cell_image.find(".//a:blip", ns)
                if c_nv_pr is None or blip is None:
                    continue

                image_name = c_nv_pr.attrib.get("name")
                embed_id = blip.attrib.get(f'{{{ns["r"]}}}embed')
                if image_name and embed_id:
                    cell_image_embed_to_name[embed_id] = image_name

            with self.zf.open(rels_path) as f:
                rel_root = ET.parse(f).getroot()

            rel_ns = {
                "pr": "http://schemas.openxmlformats.org/package/2006/relationships"
            }
            for rel in rel_root.findall("pr:Relationship", rel_ns):
                rel_id = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                if rel_id and target:
                    image_name = cell_image_embed_to_name.get(rel_id)
                    if not image_name:
                        logger.warning(
                            f"跳过缺少 cellImage 名称映射的关系: {rel_id}"
                        )
                        continue
                    self.cell_image_map[image_name] = target

        except Exception as e:
            logger.warning(f"解析 cellimages 映射失败: {e}")
            return {}

        return self.cell_image_map

    @staticmethod
    def _escape_text_with_line_breaks(text: str) -> str:
        return (
            html.escape(text)
            .replace("\r\n", "\n")
            .replace("\r", "\n")
            .replace("\n", "<br>")
        )

    @staticmethod
    def _get_cell_hyperlink_target(cell) -> str:
        hyperlink = getattr(cell, "hyperlink", None)
        if not hyperlink:
            return ""

        target = getattr(hyperlink, "target", None)
        if target:
            return str(target)

        location = getattr(hyperlink, "location", None)
        if location:
            return f"#{location}"

        return ""

    @staticmethod
    def _sanitize_hyperlink_target(target: str) -> str:
        href = target.strip()
        if not href:
            return ""

        if href.lower().startswith(("javascript:", "data:", "vbscript:")):
            return ""

        parsed = urlparse(href)
        allowed_schemes = {"http", "https", "mailto", "ftp"}
        scheme = parsed.scheme.lower() if parsed.scheme else ""
        if scheme and scheme not in allowed_schemes:
            return ""

        return html.escape(href, quote=True)

    @staticmethod
    def _apply_inline_font_tags(text_html: str, inline_font) -> str:
        if not text_html or inline_font is None:
            return text_html

        wrapped = text_html
        vert_align = getattr(inline_font, "vertAlign", None)
        if vert_align == "superscript":
            wrapped = f"<sup>{wrapped}</sup>"
        elif vert_align == "subscript":
            wrapped = f"<sub>{wrapped}</sub>"

        if getattr(inline_font, "strike", False):
            wrapped = f"<s>{wrapped}</s>"
        if getattr(inline_font, "u", None):
            wrapped = f"<u>{wrapped}</u>"
        if getattr(inline_font, "i", False):
            wrapped = f"<em>{wrapped}</em>"
        if getattr(inline_font, "b", False):
            wrapped = f"<strong>{wrapped}</strong>"

        return wrapped

    @staticmethod
    def _contains_block_level_html(content: str) -> bool:
        return bool(
            re.search(
                r"<\s*(p|ul|ol|li|div|table|blockquote|pre|h[1-6])\b",
                content,
                re.IGNORECASE,
            )
        )

    def _render_cell_inner_html(self, content: str, is_html: bool) -> str:
        if not content:
            return "<p></p>"

        if is_html and self._contains_block_level_html(content):
            return content

        return f"<p>{content}</p>"

    def _cell_value_to_html(self, cell) -> tuple[str, bool]:
        if cell.value is None:
            return "", False

        link_target = self._sanitize_hyperlink_target(
            self._get_cell_hyperlink_target(cell)
        )

        if isinstance(cell.value, CellRichText):
            html_parts = []
            for part in cell.value:
                if hasattr(part, "text"):
                    part_text = self._escape_text_with_line_breaks(
                        str(getattr(part, "text", ""))
                    )
                    html_parts.append(
                        self._apply_inline_font_tags(
                            part_text,
                            getattr(part, "font", None),
                        )
                    )
                else:
                    html_parts.append(self._escape_text_with_line_breaks(str(part)))

            rich_text_html = "".join(html_parts)
            if link_target and rich_text_html:
                rich_text_html = f'<a href="{link_target}">{rich_text_html}</a>'
            return rich_text_html, True

        plain_text = str(cell.value)
        if link_target and plain_text:
            escaped_text = self._escape_text_with_line_breaks(plain_text)
            return f'<a href="{link_target}">{escaped_text}</a>', True

        return plain_text, False

    def _extract_cell_style(self, cell):
        """Extract styles from an openpyxl cell."""
        style = {}
        if cell.font:
            if cell.font.b:
                style["font-weight"] = "bold"
            if cell.font.i:
                style["font-style"] = "italic"
            if cell.font.u:
                style["text-decoration"] = "underline"
            if cell.font.strike:
                style["text-decoration"] = "line-through"
            if (
                cell.font.color
                and hasattr(cell.font.color, "rgb")
                and cell.font.color.rgb
            ):
                # Color might be ARGB "FF000000"
                color = cell.font.color.rgb
                if isinstance(color, str) and len(color) == 8:
                    style["color"] = "#" + color[2:]
                elif isinstance(color, str):
                    style["color"] = "#" + color

        if cell.alignment:
            if cell.alignment.horizontal:
                style["text-align"] = cell.alignment.horizontal
            if cell.alignment.vertical:
                style["vertical-align"] = cell.alignment.vertical

        if cell.fill and cell.fill.patternType == "solid" and cell.fill.fgColor:
            # handle bg color
            color = cell.fill.fgColor.rgb
            if (
                hasattr(cell.fill.fgColor, "type")
                and cell.fill.fgColor.type == "rgb"
                and color
            ):
                if isinstance(color, str) and len(color) == 8:
                    style["background-color"] = "#" + color[2:]
        return style

    @staticmethod
    def _get_sheet_content_layer(sheet: Worksheet):
        """根据工作表的可见性返回对应的内容层。

        若工作表可见，返回 None（默认层）；否则返回 INVISIBLE 层。

        参数：
            sheet: 待检查的工作表。

        返回：
            ContentLayer.INVISIBLE 或 None。
        """
        return (
            None if sheet.sheet_state == Worksheet.SHEETSTATE_VISIBLE else "INVISIBLE"
        )
