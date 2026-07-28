# Copyright (c) Opendatalab. All rights reserved.
import base64
import math
from dataclasses import dataclass, field
from datetime import date, datetime, time
from html import escape
from io import BytesIO
from typing import Any, Final

from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils.cell import range_to_tuple
from openpyxl.utils.datetime import MAC_EPOCH, WINDOWS_EPOCH, from_excel
from reportlab.graphics import renderSVG
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.shapes import Drawing, Group, Line, Rect, String
from reportlab.lib import colors


_CHART_NS: Final = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_DRAWING_NS: Final = "http://schemas.openxmlformats.org/drawingml/2006/main"
_NS: Final = {"c": _CHART_NS, "a": _DRAWING_NS}
_MAX_CACHE_INDEX_SPAN: Final = 100_000
_DEFAULT_CHART_WIDTH: Final = 960
_DEFAULT_CHART_HEIGHT: Final = 540
_CHART_PALETTE: Final = (
    "4472C4",
    "ED7D31",
    "70AD47",
    "A5A5A5",
    "FFC000",
    "5B9BD5",
    "C00000",
    "00B050",
)
_SCHEME_COLORS: Final = {
    "dk1": "000000",
    "lt1": "FFFFFF",
    "dk2": "1F497D",
    "lt2": "EEECE1",
    "accent1": "4F81BD",
    "accent2": "C0504D",
    "accent3": "9BBB59",
    "accent4": "8064A2",
    "accent5": "4BACC6",
    "accent6": "F79646",
}
_PLOT_TAGS: Final = (
    "areaChart",
    "area3DChart",
    "barChart",
    "bar3DChart",
    "bubbleChart",
    "doughnutChart",
    "lineChart",
    "line3DChart",
    "ofPieChart",
    "pieChart",
    "pie3DChart",
    "radarChart",
    "scatterChart",
    "stockChart",
    "surfaceChart",
    "surface3DChart",
)


@dataclass
class SeriesSpec:
    name_formula: str | None = None
    literal_name: str | None = None
    cat_formula: str | None = None
    x_formula: str | None = None
    val_formula: str | None = None
    y_formula: str | None = None
    bubble_size_formula: str | None = None
    cached_categories: list[str] = field(default_factory=list)
    cached_x_values: list[str] = field(default_factory=list)
    cached_values: list[str] = field(default_factory=list)
    cached_bubble_sizes: list[str] = field(default_factory=list)
    line_color: str | None = None
    line_dash: str | None = None
    line_width: float | None = None


@dataclass
class ChartSpec:
    chart_type: str
    plot_kind: str
    title: str = ""
    category_axis_title: str = ""
    value_axis_title: str = ""
    x_axis_title: str = ""
    has_date_axis: bool = False
    date_1904: bool = False
    series: list[SeriesSpec] = field(default_factory=list)
    x_axis_min: float | None = None
    x_axis_max: float | None = None
    x_axis_major_unit: float | None = None
    y_axis_min: float | None = None
    y_axis_max: float | None = None
    y_axis_major_unit: float | None = None


def html_table_from_excel_bytes(excel_bytes: bytes) -> str:
    """Convert the first non-empty worksheet in an embedded workbook to HTML."""
    if not excel_bytes:
        return ""

    try:
        workbook = load_workbook(
            filename=BytesIO(excel_bytes),
            data_only=True,
            read_only=True,
        )
    except Exception:
        return ""

    try:
        for worksheet in workbook.worksheets:
            rows = _extract_non_empty_worksheet_rows(worksheet)
            if rows:
                return _render_embedded_workbook_table(rows)
    finally:
        workbook.close()

    return ""


def _extract_non_empty_worksheet_rows(worksheet) -> list[list[str]]:
    """提取工作表中首尾有内容的行，避免空 sheet 或尾部空列撑大兜底表格。"""
    raw_rows: list[list[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        stringified_row = [_stringify_cell_value(value) for value in row]
        raw_rows.append(_trim_trailing_empty_values(stringified_row))

    first_row_idx = _find_first_non_empty_row(raw_rows)
    if first_row_idx is None:
        return []

    last_row_idx = _find_last_non_empty_row(raw_rows)
    if last_row_idx is None:
        return []

    last_row_exclusive = last_row_idx + 1
    rows = raw_rows[first_row_idx:last_row_exclusive]
    width = max(len(row) for row in rows)
    return [row + [""] * (width - len(row)) for row in rows]


def _trim_trailing_empty_values(values: list[str]) -> list[str]:
    """移除行尾空值，保留中间空单元格的位置。"""
    end = len(values)
    while end > 0 and values[end - 1] == "":
        end -= 1
    return values[:end]


def _find_first_non_empty_row(rows: list[list[str]]) -> int | None:
    """返回第一行非空行的索引，未找到时返回 None。"""
    for idx, row in enumerate(rows):
        if any(value != "" for value in row):
            return idx
    return None


def _find_last_non_empty_row(rows: list[list[str]]) -> int | None:
    """返回最后一行非空行的索引。"""
    for idx in range(len(rows) - 1, -1, -1):
        if any(value != "" for value in rows[idx]):
            return idx
    return None


def _render_embedded_workbook_table(rows: list[list[str]]) -> str:
    """将嵌入 workbook 的二维数据渲染为紧凑 HTML 表格，首行作为表头。"""
    if not rows:
        return ""

    headers = rows[0]
    data_rows = rows[1:]
    html_parts = ["<table><thead><tr>"]
    for header in headers:
        html_parts.append(f"<th>{escape(header)}</th>")
    html_parts.append("</tr></thead><tbody>")

    for row in data_rows:
        html_parts.append("<tr>")
        for value in row:
            html_parts.append(f"<td>{escape(value)}</td>")
        html_parts.append("</tr>")

    html_parts.append("</tbody></table>")
    return "".join(html_parts)


def extract_chart_html_from_ooxml(chart_xml: bytes, workbook_bytes: bytes | None) -> str:
    spec = parse_chart_spec_from_ooxml(chart_xml)
    if spec is None or not spec.series:
        if workbook_bytes:
            return html_table_from_excel_bytes(workbook_bytes)
        return ""

    chart_cache_html = render_chart_html_from_cache(spec)

    if workbook_bytes:
        chart_html = render_chart_html_from_workbook(spec, workbook_bytes)
        if chart_html:
            return chart_html
        if chart_cache_html:
            return chart_cache_html
        workbook_table_html = html_table_from_excel_bytes(workbook_bytes)
        if workbook_table_html:
            return workbook_table_html

    return chart_cache_html


def render_chart_svg_from_ooxml(
    chart_xml: bytes,
    width: int = _DEFAULT_CHART_WIDTH,
    height: int = _DEFAULT_CHART_HEIGHT,
) -> str:
    """Render supported OOXML charts as an SVG data URI using cached display data."""
    spec = parse_chart_spec_from_ooxml(chart_xml)
    if spec is None or not spec.series:
        return ""

    drawing = _build_chart_drawing(spec, width, height)
    if drawing is None:
        return ""

    svg = renderSVG.drawToString(drawing)
    encoded = base64.b64encode(svg.encode("utf-8")).decode("ascii")
    return f"data:image/svg+xml;base64,{encoded}"


def _build_chart_drawing(spec: ChartSpec, width: int, height: int) -> Drawing | None:
    width = max(int(width), 320)
    height = max(int(height), 240)
    if spec.plot_kind == "scatter":
        return _build_scatter_chart_drawing(spec, width, height)
    if spec.chart_type in {"barChart", "bar3DChart"}:
        return _build_bar_chart_drawing(spec, width, height)
    return None


def _chart_number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _scatter_series_data(
    spec: ChartSpec,
) -> list[tuple[SeriesSpec, str, list[tuple[float, float]]]]:
    series_data = []
    for index, series in enumerate(spec.series, start=1):
        points = []
        for raw_x, raw_y in zip(series.cached_x_values, series.cached_values):
            x_value = _chart_number(raw_x)
            y_value = _chart_number(raw_y)
            if x_value is not None and y_value is not None:
                points.append((x_value, y_value))
        if points:
            series_data.append((series, _resolve_series_name(series, index), points))
    return series_data


def _category_series_data(
    spec: ChartSpec,
) -> tuple[list[str], list[tuple[SeriesSpec, str, list[float]]]]:
    categories = next(
        (series.cached_categories for series in spec.series if series.cached_categories),
        [],
    )
    row_count = max(
        len(categories),
        max((len(series.cached_values) for series in spec.series), default=0),
    )
    if row_count == 0:
        return [], []
    if not categories:
        categories = [str(index) for index in range(1, row_count + 1)]
    elif len(categories) < row_count:
        categories = [*categories, *[""] * (row_count - len(categories))]

    series_data = []
    for index, series in enumerate(spec.series, start=1):
        values = []
        for raw_value in series.cached_values:
            value = _chart_number(raw_value)
            values.append(value if value is not None else 0.0)
        values.extend([0.0] * (row_count - len(values)))
        series_data.append((series, _resolve_series_name(series, index), values))
    return categories, series_data


def _series_color(series: SeriesSpec, index: int):
    color_value = series.line_color or _CHART_PALETTE[index % len(_CHART_PALETTE)]
    return colors.HexColor(f"#{color_value}")


def _line_dash_array(dash_name: str | None) -> list[int] | None:
    if dash_name in {"dash", "sysDash", "lgDash"}:
        return [8, 5]
    if dash_name in {"dot", "sysDot"}:
        return [2, 4]
    if dash_name in {"dashDot", "sysDashDot", "lgDashDot"}:
        return [8, 4, 2, 4]
    if dash_name in {"lgDashDotDot", "sysDashDotDot"}:
        return [10, 4, 2, 4, 2, 4]
    return None


def _configure_value_axis(axis, value_min, value_max, major_unit) -> None:
    axis.visibleGrid = True
    axis.gridStrokeColor = colors.HexColor("#D9D9D9")
    axis.gridStrokeWidth = 0.6
    axis.gridStrokeDashArray = [1, 2]
    axis.strokeColor = colors.HexColor("#A6A6A6")
    axis.labels.fillColor = colors.HexColor("#262626")
    axis.labels.fontSize = 10
    if value_min is not None:
        axis.valueMin = value_min
    if value_max is not None:
        axis.valueMax = value_max
    if major_unit is not None and major_unit > 0:
        axis.valueStep = major_unit


def _add_chart_title_and_legend(
    drawing: Drawing,
    spec: ChartSpec,
    series_info: list[tuple[SeriesSpec, str]],
    width: int,
    height: int,
    *,
    bar_swatch: bool = False,
) -> float:
    drawing.add(
        String(
            width / 2,
            height - 27,
            spec.title,
            textAnchor="middle",
            fontName="Helvetica",
            fontSize=17,
            fillColor=colors.black,
        )
    )
    if not series_info:
        return height - 55

    columns = min(3, max(1, math.ceil(len(series_info) / 2)))
    rows = math.ceil(len(series_info) / columns)
    item_width = (width - 140) / columns
    for index, (series, name) in enumerate(series_info):
        column = index % columns
        row = index // columns
        x = 70 + column * item_width
        y = height - 57 - row * 19
        color = _series_color(series, index)
        if bar_swatch:
            drawing.add(Rect(x, y - 4, 24, 9, fillColor=color, strokeColor=color))
        else:
            legend_line = Line(
                x,
                y,
                x + 28,
                y,
                strokeColor=color,
                strokeWidth=series.line_width or 2,
            )
            dash_array = _line_dash_array(series.line_dash)
            if dash_array:
                legend_line.strokeDashArray = dash_array
            drawing.add(legend_line)
        drawing.add(
            String(
                x + 35,
                y - 4,
                name,
                fontName="Helvetica",
                fontSize=10,
                fillColor=colors.HexColor("#262626"),
            )
        )
    return height - 68 - (rows - 1) * 19


def _add_axis_titles(
    drawing: Drawing,
    spec: ChartSpec,
    chart_left: float,
    chart_bottom: float,
    chart_width: float,
    chart_height: float,
) -> None:
    if spec.x_axis_title:
        drawing.add(
            String(
                chart_left + chart_width / 2,
                22,
                spec.x_axis_title,
                textAnchor="middle",
                fontName="Helvetica",
                fontSize=12,
            )
        )
    if spec.value_axis_title:
        axis_title_group = Group()
        axis_title_group.add(
            String(
                0,
                0,
                spec.value_axis_title,
                textAnchor="middle",
                fontName="Helvetica",
                fontSize=12,
            )
        )
        axis_title_group.transform = (
            0,
            1,
            -1,
            0,
            24,
            chart_bottom + chart_height / 2,
        )
        drawing.add(axis_title_group)


def _build_scatter_chart_drawing(spec: ChartSpec, width: int, height: int) -> Drawing | None:
    series_data = _scatter_series_data(spec)
    if not series_data:
        return None

    drawing = Drawing(width, height)
    drawing.add(Rect(0, 0, width, height, fillColor=colors.white, strokeColor=None))
    series_info = [(series, name) for series, name, _ in series_data]
    chart_top = _add_chart_title_and_legend(drawing, spec, series_info, width, height)
    chart_left = 72
    chart_bottom = 58
    chart_width = width - chart_left - 28
    chart_height = max(chart_top - chart_bottom, 100)

    chart = LinePlot()
    chart.x = chart_left
    chart.y = chart_bottom
    chart.width = chart_width
    chart.height = chart_height
    chart.data = [points for _, _, points in series_data]
    chart.joinedLines = True
    _configure_value_axis(
        chart.xValueAxis,
        spec.x_axis_min,
        spec.x_axis_max,
        spec.x_axis_major_unit,
    )
    _configure_value_axis(
        chart.yValueAxis,
        spec.y_axis_min,
        spec.y_axis_max,
        spec.y_axis_major_unit,
    )
    for index, (series, _, _) in enumerate(series_data):
        chart.lines[index].strokeColor = _series_color(series, index)
        chart.lines[index].strokeWidth = series.line_width or 2
        dash_array = _line_dash_array(series.line_dash)
        if dash_array:
            chart.lines[index].strokeDashArray = dash_array

    drawing.add(chart)
    _add_axis_titles(
        drawing,
        spec,
        chart_left,
        chart_bottom,
        chart_width,
        chart_height,
    )
    return drawing


def _build_bar_chart_drawing(spec: ChartSpec, width: int, height: int) -> Drawing | None:
    categories, series_data = _category_series_data(spec)
    if not series_data:
        return None

    drawing = Drawing(width, height)
    drawing.add(Rect(0, 0, width, height, fillColor=colors.white, strokeColor=None))
    series_info = [(series, name) for series, name, _ in series_data]
    chart_top = _add_chart_title_and_legend(
        drawing,
        spec,
        series_info,
        width,
        height,
        bar_swatch=True,
    )
    chart_left = 72
    chart_bottom = 58
    chart_width = width - chart_left - 28
    chart_height = max(chart_top - chart_bottom, 100)

    chart = VerticalBarChart()
    chart.x = chart_left
    chart.y = chart_bottom
    chart.width = chart_width
    chart.height = chart_height
    chart.data = [values for _, _, values in series_data]
    chart.categoryAxis.categoryNames = categories
    chart.categoryAxis.labels.fontSize = 9
    chart.categoryAxis.labels.fillColor = colors.HexColor("#262626")
    chart.categoryAxis.strokeColor = colors.HexColor("#A6A6A6")
    _configure_value_axis(
        chart.valueAxis,
        spec.y_axis_min,
        spec.y_axis_max,
        spec.y_axis_major_unit,
    )
    for index, (series, _, _) in enumerate(series_data):
        color = _series_color(series, index)
        chart.bars[index].fillColor = color
        chart.bars[index].strokeColor = color

    drawing.add(chart)
    _add_axis_titles(
        drawing,
        spec,
        chart_left,
        chart_bottom,
        chart_width,
        chart_height,
    )
    return drawing


def parse_chart_spec_from_ooxml(chart_xml: bytes) -> ChartSpec | None:
    try:
        root = etree.fromstring(
            chart_xml,
            parser=etree.XMLParser(
                load_dtd=False,
                resolve_entities=False,
                no_network=True,
            ),
        )
    except (etree.XMLSyntaxError, TypeError, ValueError):
        return None

    plot_area = root.find(".//c:plotArea", namespaces=_NS)
    if plot_area is None:
        return None

    plot_elements = _collect_plot_elements(plot_area)
    if not plot_elements:
        return None

    has_date_axis = plot_area.find("c:dateAx", namespaces=_NS) is not None
    plot_kinds = {
        _plot_kind_from_tag_name(tag_name, has_date_axis)
        for tag_name, _ in plot_elements
    }
    if plot_kinds == {"scatter"}:
        plot_kind = "scatter"
    elif plot_kinds == {"bubble"}:
        plot_kind = "bubble"
    elif plot_kinds <= {"category", "date"}:
        plot_kind = "date" if "date" in plot_kinds else "category"
    else:
        return None

    category_axis_title = ""
    axis = plot_area.find("c:dateAx", namespaces=_NS)
    if axis is None:
        axis = plot_area.find("c:catAx", namespaces=_NS)
    if axis is not None:
        category_axis_title = _extract_title_text(axis.find("c:title", namespaces=_NS))

    x_axis_title = ""
    value_axis_title = ""
    x_axis_min = x_axis_max = x_axis_major_unit = None
    y_axis_min = y_axis_max = y_axis_major_unit = None
    if plot_kind in {"scatter", "bubble"}:
        x_axis_found = False
        y_axis_found = False
        for axis in plot_area.findall("c:valAx", namespaces=_NS):
            axis_pos = axis.find("c:axPos", namespaces=_NS)
            axis_position = axis_pos.get("val") if axis_pos is not None else ""
            title = _extract_title_text(axis.find("c:title", namespaces=_NS))
            axis_min, axis_max, major_unit = _extract_axis_scale(axis)
            if axis_position == "b" and not x_axis_found:
                x_axis_title = title
                x_axis_min = axis_min
                x_axis_max = axis_max
                x_axis_major_unit = major_unit
                x_axis_found = True
            elif axis_position == "l" and not y_axis_found:
                value_axis_title = title
                y_axis_min = axis_min
                y_axis_max = axis_max
                y_axis_major_unit = major_unit
                y_axis_found = True
        if not x_axis_title:
            x_axis_title = category_axis_title
    else:
        axis = plot_area.find("c:valAx", namespaces=_NS)
        if axis is not None:
            value_axis_title = _extract_title_text(axis.find("c:title", namespaces=_NS))
            y_axis_min, y_axis_max, y_axis_major_unit = _extract_axis_scale(axis)

    series_specs = []
    for _, plot_element in plot_elements:
        for series_element in plot_element.findall("c:ser", namespaces=_NS):
            series_specs.append(
                SeriesSpec(
                    name_formula=_extract_tx_formula(series_element.find("c:tx", namespaces=_NS)),
                    literal_name=_extract_tx_text(series_element.find("c:tx", namespaces=_NS)),
                    cat_formula=_extract_reference_formula(series_element.find("c:cat", namespaces=_NS)),
                    x_formula=_extract_reference_formula(series_element.find("c:xVal", namespaces=_NS)),
                    val_formula=_extract_reference_formula(series_element.find("c:val", namespaces=_NS)),
                    y_formula=_extract_reference_formula(series_element.find("c:yVal", namespaces=_NS)),
                    bubble_size_formula=_extract_reference_formula(
                        series_element.find("c:bubbleSize", namespaces=_NS)
                    ),
                    cached_categories=_extract_reference_cache(
                        series_element.find("c:cat", namespaces=_NS),
                        date_hint=has_date_axis,
                        date_1904=_chart_uses_date_1904(root),
                    ),
                    cached_x_values=_extract_reference_cache(
                        series_element.find("c:xVal", namespaces=_NS)
                    ),
                    cached_values=_extract_reference_cache(
                        _first_non_none(
                            series_element.find("c:val", namespaces=_NS),
                            series_element.find("c:yVal", namespaces=_NS),
                        )
                    ),
                    cached_bubble_sizes=_extract_reference_cache(
                        series_element.find("c:bubbleSize", namespaces=_NS)
                    ),
                    line_color=_extract_series_line_color(series_element),
                    line_dash=_extract_series_line_dash(series_element),
                    line_width=_extract_series_line_width(series_element),
                )
            )

    return ChartSpec(
        chart_type=(
            plot_elements[0][0]
            if len(plot_elements) == 1
            else "comboChart"
        ),
        plot_kind=plot_kind,
        title=_extract_title_text(root.find(".//c:chart/c:title", namespaces=_NS)),
        category_axis_title=category_axis_title,
        value_axis_title=value_axis_title,
        x_axis_title=x_axis_title,
        has_date_axis=has_date_axis,
        date_1904=_chart_uses_date_1904(root),
        series=series_specs,
        x_axis_min=x_axis_min,
        x_axis_max=x_axis_max,
        x_axis_major_unit=x_axis_major_unit,
        y_axis_min=y_axis_min,
        y_axis_max=y_axis_max,
        y_axis_major_unit=y_axis_major_unit,
    )


def render_chart_html_from_workbook(spec: ChartSpec, workbook_bytes: bytes) -> str:
    try:
        workbook = load_workbook(
            filename=BytesIO(workbook_bytes),
            data_only=True,
            read_only=True,
        )
    except Exception:
        return ""

    try:
        if spec.plot_kind in {"category", "date"}:
            return _render_category_like_chart_from_workbook(spec, workbook)
        if spec.plot_kind == "scatter":
            return _render_scatter_like_chart_from_workbook(spec, workbook)
        if spec.plot_kind == "bubble":
            return _render_bubble_chart_from_workbook(spec, workbook)
        return ""
    finally:
        workbook.close()


def render_chart_html_from_cache(spec: ChartSpec) -> str:
    if spec.plot_kind in {"category", "date"}:
        categories = []
        for series in spec.series:
            if series.cached_categories:
                categories = series.cached_categories
                break

        series_names = []
        series_values = []
        for idx, series in enumerate(spec.series, start=1):
            series_names.append(_resolve_series_name(series, idx))
            series_values.append(series.cached_values)

        row_count = max(
            len(categories),
            max((len(values) for values in series_values), default=0),
        )
        if not series_names or row_count == 0:
            return ""

        headers = [spec.category_axis_title or ""] + series_names
        columns = [categories] + series_values
        return _render_html_table(headers, columns, row_count)

    if spec.plot_kind == "scatter":
        return _render_scatter_like_chart_from_cache(spec)

    if spec.plot_kind == "bubble":
        return _render_bubble_chart_from_cache(spec)

    return ""


def _render_category_like_chart_from_workbook(spec: ChartSpec, workbook) -> str:
    categories = []

    for series in spec.series:
        if not series.cat_formula:
            continue
        read_result = _read_formula_vector(workbook, series.cat_formula)
        if read_result is None:
            return ""
        _, values = read_result
        categories = values
        break

    series_names = []
    series_values = []
    for idx, series in enumerate(spec.series, start=1):
        if not series.val_formula:
            return ""
        read_result = _read_formula_vector(workbook, series.val_formula)
        if read_result is None:
            return ""
        _, values = read_result
        series_names.append(_resolve_series_name(series, idx, workbook))
        series_values.append(values)

    row_count = max(
        len(categories),
        max((len(values) for values in series_values), default=0),
    )
    if not series_names or row_count == 0:
        return ""

    headers = [spec.category_axis_title or ""] + series_names
    columns = [
        _stringify_series_values(
            categories,
            date_hint=spec.has_date_axis,
            date_1904=spec.date_1904,
        )
    ]
    columns.extend(_stringify_series_values(values) for values in series_values)
    return _render_html_table(headers, columns, row_count)


def _render_scatter_like_chart_from_workbook(spec: ChartSpec, workbook) -> str:
    x_sequences, series_names, series_y_values = _read_scatter_axes_from_workbook(
        spec,
        workbook,
    )
    return _render_scatter_like_chart_table(
        x_sequences,
        series_names,
        series_y_values,
        x_axis_title=spec.x_axis_title,
    )


def _render_bubble_chart_from_workbook(spec: ChartSpec, workbook) -> str:
    x_sequences, series_names, series_y_values, series_sizes = _read_bubble_axes_from_workbook(
        spec,
        workbook,
    )
    return _render_bubble_chart_table(
        x_sequences,
        series_names,
        series_y_values,
        series_sizes,
        x_axis_title=spec.x_axis_title,
    )


def _render_scatter_like_chart_from_cache(spec: ChartSpec) -> str:
    x_sequences = []
    series_names = []
    series_y_values = []
    for idx, series in enumerate(spec.series, start=1):
        if not series.cached_x_values or not series.cached_values:
            return ""
        x_sequences.append(series.cached_x_values)
        series_names.append(_resolve_series_name(series, idx))
        series_y_values.append(series.cached_values)

    return _render_scatter_like_chart_table(
        x_sequences,
        series_names,
        series_y_values,
        x_axis_title=spec.x_axis_title,
    )


def _render_bubble_chart_from_cache(spec: ChartSpec) -> str:
    x_sequences = []
    series_names = []
    series_y_values = []
    series_sizes = []
    for idx, series in enumerate(spec.series, start=1):
        if not series.cached_x_values or not series.cached_values or not series.cached_bubble_sizes:
            return ""
        x_sequences.append(series.cached_x_values)
        series_names.append(_resolve_series_name(series, idx))
        series_y_values.append(series.cached_values)
        series_sizes.append(series.cached_bubble_sizes)

    return _render_bubble_chart_table(
        x_sequences,
        series_names,
        series_y_values,
        series_sizes,
        x_axis_title=spec.x_axis_title,
    )


def _read_scatter_axes_from_workbook(spec: ChartSpec, workbook):
    x_sequences = []
    series_names = []
    series_y_values = []

    for idx, series in enumerate(spec.series, start=1):
        if not series.x_formula or not series.y_formula:
            return None, [], []

        x_read = _read_formula_vector(workbook, series.x_formula)
        if x_read is None:
            return None, [], []
        _, x_values = x_read
        x_sequences.append(x_values)

        y_read = _read_formula_vector(workbook, series.y_formula)
        if y_read is None:
            return None, [], []
        _, y_values = y_read
        series_names.append(_resolve_series_name(series, idx, workbook))
        series_y_values.append(y_values)

    return x_sequences, series_names, series_y_values


def _read_bubble_axes_from_workbook(spec: ChartSpec, workbook):
    x_sequences = []
    series_names = []
    series_y_values = []
    series_sizes = []

    for idx, series in enumerate(spec.series, start=1):
        if not series.x_formula or not series.y_formula or not series.bubble_size_formula:
            return None, [], [], []

        x_read = _read_formula_vector(workbook, series.x_formula)
        if x_read is None:
            return None, [], [], []
        _, x_values = x_read
        x_sequences.append(x_values)

        y_read = _read_formula_vector(workbook, series.y_formula)
        bubble_size_read = _read_formula_vector(workbook, series.bubble_size_formula)
        if y_read is None or bubble_size_read is None:
            return None, [], [], []

        series_names.append(_resolve_series_name(series, idx, workbook))
        series_y_values.append(y_read[1])
        series_sizes.append(bubble_size_read[1])

    return x_sequences, series_names, series_y_values, series_sizes


def _read_formula_vector(workbook, formula: str):
    parsed = _parse_formula(formula)
    if parsed is None:
        return None

    sheet_name, min_col, min_row, max_col, max_row = parsed

    if min_col != max_col and min_row != max_row:
        return None

    try:
        worksheet = workbook[sheet_name]
    except KeyError:
        return None

    rows = worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    )
    if min_col == max_col:
        values = [row[0] for row in rows]
    else:
        values = list(next(rows, ()))

    return sheet_name, _trim_trailing_empty_series_values(values)


def _trim_trailing_empty_series_values(values: list[Any]) -> list[Any]:
    end = len(values)
    while end > 0 and values[end - 1] in (None, ""):
        end -= 1
    return values[:end]


def _read_formula_scalar(workbook, formula: str) -> str | None:
    read_result = _read_formula_vector(workbook, formula)
    if read_result is None:
        return None

    _, values = read_result
    if not values:
        return None

    value = values[0]
    if value in (None, ""):
        return None
    return _stringify_cell_value(value)


def _parse_formula(formula: str):
    formula = formula.strip()
    if not formula:
        return None
    if formula.startswith("="):
        formula = formula[1:]

    try:
        sheet_name, bounds = range_to_tuple(formula)
    except ValueError:
        return None

    if not all(isinstance(bound, int) for bound in bounds):
        return None

    min_col, min_row, max_col, max_row = bounds
    return _unescape_formula_sheet_name(sheet_name), min_col, min_row, max_col, max_row


def _unescape_formula_sheet_name(sheet_name: str) -> str:
    return sheet_name.replace("''", "'")


def _extract_reference_formula(container) -> str | None:
    ref_element = _find_reference_element(container)
    if ref_element is None:
        return None
    formula_element = ref_element.find("c:f", namespaces=_NS)
    if formula_element is None or formula_element.text is None:
        return None
    return formula_element.text.strip()


def _extract_reference_cache(
    container,
    *,
    date_hint: bool = False,
    date_1904: bool = False,
) -> list[str]:
    ref_element = _find_reference_element(container)
    if ref_element is None:
        return []

    tag_name = etree.QName(ref_element).localname
    if tag_name == "multiLvlStrRef":
        return _extract_multilevel_string_cache(ref_element)

    cache_element = ref_element.find("c:strCache", namespaces=_NS)
    if cache_element is None:
        cache_element = ref_element.find("c:numCache", namespaces=_NS)
    if cache_element is None:
        return []

    return _extract_cache_points(
        cache_element,
        date_hint=date_hint,
        date_1904=date_1904,
    )


def _extract_cache_points(
    cache_element,
    *,
    date_hint: bool = False,
    date_1904: bool = False,
) -> list[str]:
    points = {}
    for point in cache_element.findall("c:pt", namespaces=_NS):
        raw_index = point.get("idx")
        if raw_index is None:
            continue
        try:
            point_index = int(raw_index)
        except ValueError:
            continue

        value_element = point.find("c:v", namespaces=_NS)
        raw_value = value_element.text if value_element is not None else ""
        points[point_index] = _stringify_cache_value(
            raw_value,
            date_hint=date_hint,
            date_1904=date_1904,
        )

    if not points:
        return []

    max_index = max(points.keys())
    if max_index + 1 > _MAX_CACHE_INDEX_SPAN:
        return []

    return [points.get(index, "") for index in range(max_index + 1)]


def _extract_multilevel_string_cache(ref_element) -> list[str]:
    level_maps = []
    max_index = -1
    for level in ref_element.findall("c:multiLvlStrCache/c:lvl", namespaces=_NS):
        values = {}
        for point in level.findall("c:pt", namespaces=_NS):
            raw_index = point.get("idx")
            if raw_index is None:
                continue
            try:
                point_index = int(raw_index)
            except ValueError:
                continue
            value_element = point.find("c:v", namespaces=_NS)
            values[point_index] = value_element.text if value_element is not None else ""
            max_index = max(max_index, point_index)
        level_maps.append(values)

    if max_index < 0:
        return []

    if max_index + 1 > _MAX_CACHE_INDEX_SPAN:
        return []

    rows = []
    for point_index in range(max_index + 1):
        parts = [
            value_map[point_index]
            for value_map in level_maps
            if value_map.get(point_index)
        ]
        rows.append(" / ".join(parts))
    return rows


def _extract_tx_formula(tx_element) -> str | None:
    if tx_element is None:
        return None
    str_ref = tx_element.find("c:strRef", namespaces=_NS)
    if str_ref is None:
        return None
    formula_element = str_ref.find("c:f", namespaces=_NS)
    if formula_element is None or formula_element.text is None:
        return None
    return formula_element.text.strip()


def _extract_tx_text(tx_element) -> str | None:
    if tx_element is None:
        return None

    str_cache = tx_element.find("c:strRef/c:strCache", namespaces=_NS)
    if str_cache is not None:
        values = _extract_cache_points(str_cache)
        return values[0] if values else None

    value_element = tx_element.find("c:v", namespaces=_NS)
    if value_element is not None and value_element.text:
        return value_element.text.strip()

    return None


def _extract_title_text(title_element) -> str:
    if title_element is None:
        return ""
    texts = title_element.findall(".//a:t", namespaces=_NS)
    return "".join(text.text or "" for text in texts).strip()


def _float_xml_value(element) -> float | None:
    if element is None:
        return None
    try:
        value = float(element.get("val"))
    except (TypeError, ValueError):
        return None
    return value if math.isfinite(value) else None


def _extract_axis_scale(axis) -> tuple[float | None, float | None, float | None]:
    if axis is None:
        return None, None, None
    return (
        _float_xml_value(axis.find("c:scaling/c:min", namespaces=_NS)),
        _float_xml_value(axis.find("c:scaling/c:max", namespaces=_NS)),
        _float_xml_value(axis.find("c:majorUnit", namespaces=_NS)),
    )


def _extract_series_line_color(series_element) -> str | None:
    line = series_element.find("c:spPr/a:ln", namespaces=_NS)
    if line is None:
        return None
    srgb_color = line.find("a:solidFill/a:srgbClr", namespaces=_NS)
    if srgb_color is not None:
        raw_color = (srgb_color.get("val") or "").upper()
        if len(raw_color) == 6 and all(char in "0123456789ABCDEF" for char in raw_color):
            return raw_color
    scheme_color = line.find("a:solidFill/a:schemeClr", namespaces=_NS)
    if scheme_color is not None:
        return _SCHEME_COLORS.get(scheme_color.get("val") or "")
    return None


def _extract_series_line_dash(series_element) -> str | None:
    line = series_element.find("c:spPr/a:ln", namespaces=_NS)
    if line is None:
        return None
    dash = line.find("a:prstDash", namespaces=_NS)
    return dash.get("val") if dash is not None else None


def _extract_series_line_width(series_element) -> float | None:
    line = series_element.find("c:spPr/a:ln", namespaces=_NS)
    if line is None:
        return None
    try:
        width = float(line.get("w")) / 12_700.0
    except (TypeError, ValueError):
        return None
    if not math.isfinite(width) or width <= 0:
        return None
    return width


def _find_reference_element(container):
    if container is None:
        return None
    for tag_name in ("strRef", "numRef", "multiLvlStrRef"):
        ref_element = container.find(f"c:{tag_name}", namespaces=_NS)
        if ref_element is not None:
            return ref_element
    return None


def _first_non_none(*values):
    for value in values:
        if value is not None:
            return value
    return None


def _collect_plot_elements(plot_area) -> list[tuple[str, Any]]:
    plot_elements = []
    for child in plot_area:
        if not isinstance(child.tag, str):
            continue
        tag_name = etree.QName(child).localname
        if tag_name in _PLOT_TAGS:
            plot_elements.append((tag_name, child))
    return plot_elements


def _plot_kind_from_tag_name(tag_name: str, has_date_axis: bool) -> str:
    if tag_name == "scatterChart":
        return "scatter"
    if tag_name == "bubbleChart":
        return "bubble"
    if has_date_axis:
        return "date"
    return "category"


def _chart_uses_date_1904(root) -> bool:
    date_1904 = root.find("c:date1904", namespaces=_NS)
    if date_1904 is None:
        return False
    return date_1904.get("val") == "1"


def _resolve_series_name(series: SeriesSpec, index: int, workbook=None) -> str:
    if workbook is not None and series.name_formula:
        workbook_name = _read_formula_scalar(workbook, series.name_formula)
        if workbook_name:
            return workbook_name
    if series.literal_name:
        return series.literal_name
    return f"Series{index}"


def _get_shared_axis_values(sequences: list[list[Any]]) -> list[Any] | None:
    if not sequences:
        return None

    normalized = [_normalize_sequence(sequence) for sequence in sequences]
    first = normalized[0]
    if any(sequence != first for sequence in normalized[1:]):
        return None
    return sequences[0]


def _normalize_sequence(sequence: list[Any]) -> list[str]:
    return [_stringify_cell_value(value) for value in sequence]


def _render_scatter_like_chart_table(
    x_sequences: list[list[Any]] | None,
    series_names: list[str],
    series_y_values: list[list[Any]],
    *,
    x_axis_title: str,
) -> str:
    if not x_sequences or not series_names or len(x_sequences) != len(series_names):
        return ""

    shared_x_values = _get_shared_axis_values(x_sequences)
    if shared_x_values is not None:
        row_count = max(
            len(shared_x_values),
            max((len(values) for values in series_y_values), default=0),
        )
        if row_count == 0:
            return ""

        headers = [x_axis_title or ""] + series_names
        columns = [_stringify_series_values(shared_x_values)]
        columns.extend(_stringify_series_values(values) for values in series_y_values)
        return _render_html_table(headers, columns, row_count)

    headers = []
    columns = []
    row_count = 0
    for name, x_values, y_values in zip(series_names, x_sequences, series_y_values):
        headers.extend((f"{name} X", f"{name} Y"))
        columns.append(_stringify_series_values(x_values))
        columns.append(_stringify_series_values(y_values))
        row_count = max(row_count, len(x_values), len(y_values))

    if row_count == 0:
        return ""

    return _render_html_table(headers, columns, row_count)


def _render_bubble_chart_table(
    x_sequences: list[list[Any]] | None,
    series_names: list[str],
    series_y_values: list[list[Any]],
    series_sizes: list[list[Any]],
    *,
    x_axis_title: str,
) -> str:
    if (
        not x_sequences
        or not series_names
        or len(x_sequences) != len(series_names)
        or len(series_y_values) != len(series_names)
        or len(series_sizes) != len(series_names)
    ):
        return ""

    shared_x_values = _get_shared_axis_values(x_sequences)
    if shared_x_values is not None:
        row_count = max(
            len(shared_x_values),
            max((len(values) for values in series_y_values), default=0),
            max((len(values) for values in series_sizes), default=0),
        )
        if row_count == 0:
            return ""

        headers = [x_axis_title or ""]
        columns = [_stringify_series_values(shared_x_values)]
        for name, y_values, bubble_sizes in zip(series_names, series_y_values, series_sizes):
            headers.extend((name, f"{name} size"))
            columns.append(_stringify_series_values(y_values))
            columns.append(_stringify_series_values(bubble_sizes))
        return _render_html_table(headers, columns, row_count)

    headers = []
    columns = []
    row_count = 0
    for name, x_values, y_values, bubble_sizes in zip(
        series_names,
        x_sequences,
        series_y_values,
        series_sizes,
    ):
        headers.extend((f"{name} X", f"{name} Y", f"{name} size"))
        columns.append(_stringify_series_values(x_values))
        columns.append(_stringify_series_values(y_values))
        columns.append(_stringify_series_values(bubble_sizes))
        row_count = max(row_count, len(x_values), len(y_values), len(bubble_sizes))

    if row_count == 0:
        return ""

    return _render_html_table(headers, columns, row_count)


def _stringify_series_values(
    values: list[Any],
    *,
    date_hint: bool = False,
    date_1904: bool = False,
) -> list[str]:
    return [
        _stringify_cell_value(
            value,
            date_hint=date_hint,
            date_1904=date_1904,
        )
        for value in values
    ]


def _stringify_cache_value(
    value: str | None,
    *,
    date_hint: bool = False,
    date_1904: bool = False,
) -> str:
    if value in (None, ""):
        return ""

    if date_hint:
        try:
            serial = float(value)
        except (TypeError, ValueError):
            return value
        return (
            _excel_serial_to_iso(serial, date_1904=date_1904)
            or value
        )

    return value


def _stringify_cell_value(
    value: Any,
    *,
    date_hint: bool = False,
    date_1904: bool = False,
) -> str:
    if value in (None, ""):
        return ""

    if isinstance(value, datetime):
        if date_hint and value.time() == time():
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()

    if date_hint and isinstance(value, (int, float)):
        return (
            _excel_serial_to_iso(float(value), date_1904=date_1904)
            or _stringify_non_date_value(value)
        )

    return _stringify_non_date_value(value)


def _excel_serial_to_iso(serial: float, *, date_1904: bool = False) -> str | None:
    if not math.isfinite(serial):
        return None
    try:
        excel_value = from_excel(serial, MAC_EPOCH if date_1904 else WINDOWS_EPOCH)
    except (TypeError, ValueError, OverflowError):
        return None
    if isinstance(excel_value, datetime):
        if excel_value.time() == time():
            return excel_value.date().isoformat()
        return excel_value.isoformat(sep=" ")
    if isinstance(excel_value, date):
        return excel_value.isoformat()
    if isinstance(excel_value, time):
        return excel_value.isoformat()
    return str(excel_value)


def _stringify_non_date_value(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _render_html_table(headers: list[str], columns: list[list[str]], row_count: int) -> str:
    if row_count <= 0 or len(headers) != len(columns):
        return ""

    html_parts = ["<table><thead><tr>"]
    for header in headers:
        html_parts.append(f"<th>{escape(header)}</th>")
    html_parts.append("</tr></thead><tbody>")

    for row_idx in range(row_count):
        html_parts.append("<tr>")
        for column in columns:
            value = column[row_idx] if row_idx < len(column) else ""
            html_parts.append(f"<td>{escape(value)}</td>")
        html_parts.append("</tr>")

    html_parts.append("</tbody></table>")
    return "".join(html_parts)
