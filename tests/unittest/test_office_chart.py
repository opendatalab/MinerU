# Copyright (c) Opendatalab. All rights reserved.
from io import BytesIO
from typing import Any, Iterator

from openpyxl import Workbook, load_workbook

from mineru.backend.utils.office_chart import (
    ChartSpec,
    SeriesSpec,
    _read_formula_vector,
    render_chart_html_from_workbook,
)


class _StreamingWorksheet:
    def __init__(self) -> None:
        self.iter_rows_calls = 0

    def iter_rows(self, **kwargs: Any) -> Iterator[tuple[Any, ...]]:
        self.iter_rows_calls += 1
        assert kwargs == {
            "min_row": 4,
            "max_row": 50_000,
            "min_col": 1,
            "max_col": 1,
            "values_only": True,
        }
        yield (1,)
        yield (None,)
        yield (2,)
        yield (None,)

    def cell(self, **kwargs: Any) -> None:
        raise AssertionError(f"cell() must not be used for streaming worksheets: {kwargs}")


def _save_workbook(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def test_read_formula_vector_streams_range_once_and_trims_empty_tail() -> None:
    worksheet = _StreamingWorksheet()
    workbook = {"Data": worksheet}

    result = _read_formula_vector(workbook, "Data!$A$4:$A$50000")

    assert result == ("Data", [1, None, 2])
    assert worksheet.iter_rows_calls == 1


def test_read_formula_vector_preserves_internal_empty_horizontal_cells() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["B2"] = 1
    worksheet["D2"] = 3
    workbook_bytes = _save_workbook(workbook)

    read_only_workbook = load_workbook(
        BytesIO(workbook_bytes),
        data_only=True,
        read_only=True,
    )
    try:
        result = _read_formula_vector(read_only_workbook, "Data!$B$2:$Z$2")
    finally:
        read_only_workbook.close()

    assert result == ("Data", [1, None, 3])


def test_render_scatter_chart_ignores_empty_formula_tail() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["Time", "Voltage"])
    worksheet.append([1, 3.7])
    worksheet.append([2, 3.6])
    workbook_bytes = _save_workbook(workbook)
    spec = ChartSpec(
        chart_type="scatterChart",
        plot_kind="scatter",
        x_axis_title="Time",
        series=[
            SeriesSpec(
                literal_name="Voltage",
                x_formula="Data!$A$2:$A$50000",
                y_formula="Data!$B$2:$B$50000",
            )
        ],
    )

    html = render_chart_html_from_workbook(spec, workbook_bytes)

    assert html == (
        "<table><thead><tr><th>Time</th><th>Voltage</th></tr></thead><tbody>"
        "<tr><td>1</td><td>3.7</td></tr>"
        "<tr><td>2</td><td>3.6</td></tr>"
        "</tbody></table>"
    )
