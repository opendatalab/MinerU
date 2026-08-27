from __future__ import annotations

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.worksheet import Worksheet
import pytest

from mineru.model.flash.office.spreadsheet.html import render_spreadsheet_table
from mineru.model.flash.office.spreadsheet.models import AnchoredBlock, ExcelCell, ExcelTable, SheetImage
from mineru.model.flash.office.spreadsheet.projector import SpreadsheetProjector
from mineru.model.flash.office.xls.xls_converter import _XlsPageBuilder
from mineru.model.flash.office.xlsx import xlsx_converter as xlsx_converter_module
from mineru.model.flash.office.xlsx.xlsx_converter import XlsxConverter
from mineru.types import BlockType


def _cell(
    row: int,
    col: int,
    text: str,
    *,
    row_span: int = 1,
    col_span: int = 1,
    media: list[str] | None = None,
    equations: list[str] | None = None,
    text_is_html: bool = False,
    source_row: int | None = None,
    source_col: int | None = None,
) -> ExcelCell:
    """构造具有完整默认值的共享单元格测试对象。"""
    return ExcelCell(
        row=row,
        col=col,
        text=text,
        row_span=row_span,
        col_span=col_span,
        media=media or [],
        equations=equations or [],
        text_is_html=text_is_html,
        source_row=source_row,
        source_col=source_col,
    )


def test_shared_html_renderer_preserves_escape_merge_media_and_equation_order() -> None:
    """验证纯 renderer 保留转义、合并格以及文本媒体公式顺序。"""
    table = ExcelTable(
        anchor=(0, 0),
        num_rows=2,
        num_cols=2,
        data=[
            _cell(0, 0, "A&B"),
            _cell(
                0,
                1,
                "<strong>rich</strong>",
                media=['<img src="data:image/png;base64,AA==" />'],
                equations=["x+1"],
                text_is_html=True,
            ),
            _cell(1, 0, "merged", col_span=2),
        ],
    )

    rendered = render_spreadsheet_table(table)
    soup = BeautifulSoup(rendered, "html.parser")
    cells = soup.find_all(["th", "td"])

    assert "A&amp;B" in rendered
    rich_html = str(cells[1])
    assert rich_html.index("<strong>rich</strong>") < rich_html.index("<img") < rich_html.index("<eq>x+1</eq>")
    assert cells[2]["colspan"] == "2"
    assert len(soup.find_all("tr")) == 2


@pytest.mark.parametrize(
    ("media", "equations", "text_is_html"),
    [
        (["<img/>"], [], False),
        ([], ["x"], False),
        ([], [], True),
    ],
)
def test_singleton_with_structured_content_remains_table(
    media: list[str],
    equations: list[str],
    text_is_html: bool,
) -> None:
    """验证带媒体、公式或 HTML 的单格不会错误降级为文本。"""
    projector = SpreadsheetProjector()
    plain = ExcelTable(anchor=(0, 0), num_rows=1, num_cols=1, data=[_cell(0, 0, "plain")])
    structured = ExcelTable(
        anchor=(0, 0),
        num_rows=1,
        num_cols=1,
        data=[_cell(0, 0, "value", media=media, equations=equations, text_is_html=text_is_html)],
    )

    assert projector._build_block_from_excel_table(plain)["type"] == BlockType.TEXT
    assert projector._build_block_from_excel_table(structured)["type"] == BlockType.TABLE


def test_projector_materializes_safe_links_rich_text_media_and_equations() -> None:
    """验证单元格 IR 在渲染前已完整持有链接、富文本、媒体和公式。"""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet["A1"] = "link"
    sheet["A1"].hyperlink = "https://example.test/path"
    sheet["B1"] = CellRichText([TextBlock(InlineFont(b=True), "Bold"), " plain"])
    projector = SpreadsheetProjector()
    projector.workbook = workbook
    projector.math_map = {(0, 1): ["x"]}
    projector.table_image_map[(0, 1)].append("<img/>")

    link_cell = projector._build_excel_cell(sheet, 0, 0, 0, 0)
    rich_cell = projector._build_excel_cell(sheet, 0, 1, 0, 1)

    assert link_cell.text == '<a href="https://example.test/path">link</a>'
    assert link_cell.text_is_html is True
    assert rich_cell.text == "<strong>Bold</strong> plain"
    assert rich_cell.media == ["<img/>"]
    assert rich_cell.equations == ["x"]


def test_gap_discovery_sparse_cells_and_merged_spans_are_stable() -> None:
    """验证 gap 候选、稀疏扫描和合并跨度沿用当前算法。"""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet["A1"] = "left"
    sheet["C1"] = "right"
    projector = SpreadsheetProjector()
    projector.workbook = workbook

    gap_zero = projector._find_data_tables_with_gap_raw(sheet, 0)
    gap_one = projector._find_data_tables_with_gap_raw(sheet, 1)

    assert len(gap_zero) == 2
    assert [(table.num_rows, table.num_cols) for table in gap_one] == [(1, 3)]

    merged = workbook.create_sheet("Merged")
    merged["A1"] = "merged"
    merged.merge_cells("A1:B1")
    merged_table = projector._find_data_tables_with_gap_raw(merged, 0)[0]
    assert merged_table.data[0].col_span == 2

    sparse = workbook.create_sheet("Sparse")
    sparse["A1"] = "first"
    sparse.cell(row=1000, column=1000, value="last")
    initial_cell_count = len(sparse._cells)
    assert len(projector._find_data_tables_with_gap_raw(sparse, 0)) == 2
    assert len(sparse._cells) == initial_cell_count


def test_semantic_subset_filter_keeps_only_maximal_table() -> None:
    """验证候选过滤只删除语义坐标严格属于另一候选的表格。"""
    projector = SpreadsheetProjector()
    subset = ExcelTable(
        anchor=(0, 0),
        num_rows=1,
        num_cols=1,
        data=[_cell(0, 0, "a", source_row=0, source_col=0)],
    )
    superset = ExcelTable(
        anchor=(0, 0),
        num_rows=1,
        num_cols=2,
        data=[
            _cell(0, 0, "a", source_row=0, source_col=0),
            _cell(0, 1, "b", source_row=0, source_col=1),
        ],
    )

    assert projector._filter_semantic_subset_tables([subset, superset]) == [superset]


def test_hidden_sheet_titles_and_standalone_images_remain_separate() -> None:
    """验证隐藏 sheet、非空页标题和未吸收图片继续遵守原边界。"""
    workbook = Workbook()
    first = workbook.active
    assert first is not None
    first.title = "First"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = Worksheet.SHEETSTATE_HIDDEN
    workbook.create_sheet("Second")
    projector = SpreadsheetProjector()
    projector.workbook = workbook

    assert [sheet.title for sheet in projector._iter_sheets_to_convert()] == ["First", "Second"]
    sheet_pages = [("First", [{"type": BlockType.TEXT, "content": "a"}]), ("Empty", [])]
    projector._prepend_sheet_titles(sheet_pages)
    assert sheet_pages[0][1][0] == {
        "type": BlockType.PARAGRAPH_TITLE,
        "level": 2,
        "content": "First",
    }
    assert sheet_pages[1][1] == []

    projector.cur_page = []
    projector.sheet_images = [
        SheetImage(anchor=(0, 0), image_base64="used"),
        SheetImage(anchor=(1, 0), latex="x"),
        SheetImage(anchor=(2, 0), image_base64="free"),
    ]
    projector._find_images_in_sheet({(0, 0)})
    assert projector.cur_page == [{"type": BlockType.IMAGE, "image_base64": "free"}]


class _OrderingProjector(SpreadsheetProjector):
    """为稳定排序测试提供不依赖表格发现的确定性 artifacts。"""

    def _find_tables_in_sheet(
        self,
        sheet: Worksheet,
    ) -> tuple[set[tuple[int, int]], list[AnchoredBlock]]:
        """返回一个位置较后的表格 block。"""
        return set(), [((1, 0), 0, {"type": BlockType.TABLE, "content": "table"})]

    def _find_charts_in_sheet(self, sheet: Worksheet) -> list[AnchoredBlock]:
        """返回与公式同 anchor 但优先级较后的 chart。"""
        return [((0, 1), 10, {"type": BlockType.CHART, "content": "chart"})]

    def _find_additional_visual_artifacts(
        self,
        used_cells: set[tuple[int, int]],
    ) -> list[AnchoredBlock]:
        """返回与 chart 同 anchor 但优先级更高的公式。"""
        return [((0, 1), 5, {"type": BlockType.EQUATION, "content": "equation"})]

    def _find_images_in_sheet(self, used_cells: set[tuple[int, int]] | None = None) -> None:
        """在几何排序完成后追加独立图片。"""
        self.cur_page.append({"type": BlockType.IMAGE, "image_base64": "image"})


def test_sheet_projection_sorts_artifacts_before_standalone_images() -> None:
    """验证排序键固定为 anchor、priority，独立图片最后输出。"""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    projector = _OrderingProjector()
    projector.workbook = workbook

    projector._convert_sheet(sheet)

    assert [block["type"] for block in projector.cur_page] == [
        BlockType.EQUATION,
        BlockType.CHART,
        BlockType.TABLE,
        BlockType.IMAGE,
    ]


def test_xls_and_xlsx_depend_only_on_shared_projector_contract() -> None:
    """验证最终继承关系和旧 XLSX 私有表格模型均已严格迁移。"""
    assert issubclass(_XlsPageBuilder, SpreadsheetProjector)
    assert issubclass(XlsxConverter, SpreadsheetProjector)
    assert XlsxConverter not in _XlsPageBuilder.__mro__
    assert not hasattr(xlsx_converter_module, "DataRegion")
    assert not hasattr(xlsx_converter_module, "ExcelCell")
    assert not hasattr(xlsx_converter_module, "ExcelTable")
    assert not hasattr(XlsxConverter, "excel_table_to_html")
